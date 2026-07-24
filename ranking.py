import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tkinter import messagebox


# =============================
# Safe Save / Open Helpers
# =============================

def safe_save_docx(save_func, *args, **kwargs):
    """Safely save Word files and show a friendly message if file is open."""
    try:
        return save_func(*args, **kwargs)
    except PermissionError:
        messagebox.showerror(
            "File in Use",
            "❌ Cannot save the Word report.\n\nPlease close the file before trying again."
        )
        raise


def safe_save_excel(save_func, *args, **kwargs):
    """Safely save Excel files and show a friendly message if file is open."""
    try:
        return save_func(*args, **kwargs)
    except PermissionError:
        messagebox.showerror(
            "File in Use",
            "❌ Cannot save one or more Excel files.\n\n"
            "Please close all open Excel files before trying again."
        )
        raise


def safe_open_excel(path, mode="rb"):
    """Safely open Excel files and show a clear message if they’re already open."""
    try:
        return open(path, mode)
    except PermissionError:
        messagebox.showerror(
            "File in Use",
            f"❌ Cannot open '{os.path.basename(path)}'.\n\n"
            "Please close it in Excel before continuing."
        )
        raise


def format_ranking_table(file_path, top_n_to_shade=0):
    """Apply formatting:
       - Light yellow header
       - Green for Average Score & Variance
       - Light red for Ranking column
       - Black borders
       - Light green highlight ONLY for Organisation Name cells of top N applications
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.freeze_panes = "B2"

    # === Styles ===
    header_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # light yellow
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")   # light green
    purple_fill = PatternFill(start_color="E6CCFF", end_color="E6CCFF", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")   # white
    light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # light red
    header_font = Font(name="Arial", bold=True, size=11)
    normal_font = Font(name="Arial", size=11)
    wrap = Alignment(wrap_text=True, horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    # === Header ===
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap
        cell.border = thin_border

    # Identify columns
    headers = {str(cell.value).strip(): cell.column_letter for cell in ws[1] if cell.value}
    ranking_col = headers.get("Ranking")
    avg_col = headers.get("Average Score")
    var_col = headers.get("Score Range")
    name_col = headers.get("Organisation Name")

    # === Apply base formatting (white + borders) ===
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = normal_font
            cell.alignment = wrap
            cell.fill = white_fill
            cell.border = thin_border

    # === Colour entire Average Score + Marking Variance columns (green) ===
    for col_letter in [avg_col, var_col]:
        if col_letter:
            for cell in ws[col_letter]:
                if cell.row > 1:
                    cell.fill = purple_fill

    # === Ranking column (light red background) ===
    if ranking_col:
        for cell in ws[ranking_col]:
            if cell.row > 1:
                cell.fill = light_red_fill

    # === Highlight ONLY the names of the top N ranked applications ===
    if top_n_to_shade > 0 and name_col and ranking_col:
        # Collect all rank values (numeric)
        ranks = []
        for cell in ws[ranking_col]:
            if cell.row > 1 and cell.value not in (None, ""):
                try:
                    ranks.append((int(cell.value), cell.row))
                except ValueError:
                    continue

        # Sort by rank (ascending = best)
        ranks.sort(key=lambda x: x[0])

        # Highlight only the top N names
        for _, row_idx in ranks[:top_n_to_shade]:
            name_cell = ws[f"{name_col}{row_idx}"]
            name_cell.fill = green_fill

    # === Adjust widths dynamically ===
    for col_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col_cells[:200])
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_length + 3, 40)

    safe_save_excel(wb.save, file_path)




# =============================
# Main Ranking Logic
# =============================

def make_ranking(responses_path, output_ranking, scorers_folders, selected_columns,
                 top_n_to_shade=0, name_column=None):
    """
    Create final ranking Excel combining scorer scorecards + selected columns from the Responses sheet.
    """

    import re

    def clean_name(name):
        """Standardize organization names for consistent merging."""
        if pd.isna(name):
            return ""
        name = str(name).lower().strip()
        name = re.sub(r"[^a-z0-9\s]", "", name)  # remove punctuation/symbols
        name = re.sub(r"\s+", " ", name)         # collapse extra spaces
        return name

    # --- STEP 1: Load Responses safely ---
    with safe_open_excel(responses_path, "rb") as f:
        df_responses = pd.read_excel(f)
        

    # Keep only user-selected columns (plus the organisation name column)
    if selected_columns:
        keep_cols = set(selected_columns)
        if name_column:
            keep_cols.add(name_column)
        df_responses = df_responses[[c for c in df_responses.columns if c in keep_cols]]

    # --- Detect Organisation Name column ---
    if name_column and name_column in df_responses.columns:
        org_col = name_column
    else:
        org_col = next((c for c in df_responses.columns
                        if any(k in c.lower() for k in ["organisation", "group", "name"])), None)
    if not org_col:
        raise ValueError("❌ Could not identify Organisation Name column in Responses file.")

    # Detect ID column (optional)
    id_col = next((c for c in df_responses.columns if "unique" in c.lower()), None)

    # Rename consistently
    rename_map = {org_col: "Organisation Name"}
    if id_col:
        rename_map[id_col] = "Unique Application ID"
    df_responses = df_responses.rename(columns=rename_map)

    # --- Preserve original names and create cleaned version for matching ---
    df_responses["Original Organisation Name"] = df_responses["Organisation Name"]
    df_responses["Cleaned Name"] = df_responses["Organisation Name"].apply(clean_name)

    # --- STEP 2: Load all scorers' scorecards safely ---
    scorer_data = []
    for scorer_folder in scorers_folders:
        scorer_name = os.path.basename(scorer_folder).strip()
    
        # Find first Excel file in the folder
        excel_files = [
            f for f in os.listdir(scorer_folder)
            if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$")
        ]
    
        if not excel_files:
            print(f"⚠️ No Excel files found in folder: {scorer_folder}")
            continue
    
        scorecard_path = os.path.join(scorer_folder, excel_files[0])
        print(f"📘 Using scorecard: {scorecard_path}")
    
        try:
            with safe_open_excel(scorecard_path, "rb") as f:
                df = pd.read_excel(f)
        except PermissionError:
            print(f"❌ '{scorecard_path}' is open in Excel, skipping.")
            continue
        except Exception as e:
            print(f"❌ Could not read {scorecard_path}: {e}")
            continue
    
        # --- Detect and normalize columns ---
        name_col = next((c for c in df.columns
                         if any(k in c.lower() for k in ["organisation", "group", "name"])), None)
        score_col = next((c for c in df.columns
                          if any(k in c.lower() for k in ["points", "score", "mark"])), None)
        comment_col = next((c for c in df.columns
                            if any(k in c.lower() for k in ["comment", "feedback", "note"])), None)
    
        if not name_col:
            print(f"⚠️ {scorer_name}: No Organisation Name column found.")
            continue
        if not score_col:
            print(f"⚠️ {scorer_name}: No Score column found.")
            continue
    
        df = df.rename(columns={name_col: "Organisation Name", score_col: "Score"})
        if comment_col:
            df = df.rename(columns={comment_col: "Comments"})
        else:
            df["Comments"] = ""
    
        df["Cleaned Name"] = df["Organisation Name"].apply(clean_name)
        df["Scorer Name"] = scorer_name
        
    
        scorer_data.append(df)


    if not scorer_data:
        raise ValueError("❌ No valid scorecard files found in the selected folders.")

    scorer_df = pd.concat(scorer_data, ignore_index=True)
    

    # --- STEP 3: Merge scorers with responses ---
    if (
        "Unique Application ID" in df_responses.columns and
        "Unique Application ID" in scorer_df.columns
    ):
        merge_key = "Unique Application ID"
    else:
        merge_key = "Cleaned Name"

    df_rank = df_responses.drop_duplicates(subset=merge_key, keep="first")

    
    df_merged = pd.merge(df_rank, scorer_df, on=merge_key, how="left")
    


    
    

    # --- STEP 4: Clean + limit scorers ---
    df_merged["Index"] = df_merged.groupby("Cleaned Name").cumcount() + 1
    df_merged = df_merged[df_merged["Index"] <= 3]  # limit to first 3 scorers

    # --- STEP 5: Pivot into wide format ---
    pivot_df = df_merged.pivot_table(
        index=["Cleaned Name"],
        columns="Index",
        values=["Score", "Scorer Name", "Comments"],
        aggfunc="first",
        dropna=False
    )
    pivot_df.columns = [f"{col[0]}{col[1]}" for col in pivot_df.columns]
    pivot_df = pivot_df.reset_index()
    

    
    # --- STEP 6: Restore original names from Responses ---
    name_map = (
        df_responses.drop_duplicates(subset="Cleaned Name")
        .set_index("Cleaned Name")["Original Organisation Name"]
        .to_dict()
    )
    pivot_df["Organisation Name"] = pivot_df["Cleaned Name"].map(name_map).fillna(pivot_df["Cleaned Name"])

    # --- STEP 7: Add user-selected metadata columns ---
    other_cols = [c for c in df_responses.columns if c not in
                  ["Organisation Name", "Original Organisation Name", "Cleaned Name", "Unique Application ID"]]
    unique_data = df_responses.drop_duplicates(subset="Cleaned Name")[["Cleaned Name"] + other_cols]
    final_df = pd.merge(pivot_df, unique_data, on="Cleaned Name", how="left")
    final_df.drop(columns=["Cleaned Name"], inplace=True)
    

    

    # --- STEP 8: Compute averages, variance, ranking ---
    score_cols = [c for c in final_df.columns if c.startswith("Score") and not c.startswith("Scorer")]

    for col in score_cols:
        final_df[col] = pd.to_numeric(final_df[col], errors="coerce")

    if score_cols:
        final_df["Average Score"] = final_df[score_cols].mean(axis=1, skipna=True)
        final_df["Score Range"] = final_df[score_cols].max(axis=1) - final_df[score_cols].min(axis=1)
        ranked = final_df["Average Score"].replace([np.inf, -np.inf], np.nan)
        final_df["Ranking"] = ranked.rank(ascending=False, method="dense").astype("Int64")
    else:
        final_df["Average Score"] = np.nan
        final_df["Score Range"] = np.nan
        final_df["Ranking"] = np.nan
        
        

    # --- STEP 9: Reorder columns ---
    info_cols = ["Average Score", "Score Range", "Ranking"]
    base_cols = ["Organisation Name"] + info_cols
    scoring_cols = [f"{x}{i}" for i in range(1, 4) for x in ["Score", "Scorer Name", "Comments"]]
    optional_cols = [c for c in final_df.columns if c not in base_cols + scoring_cols]

    final_df = final_df[[c for c in base_cols + scoring_cols + optional_cols if c in final_df.columns]]
    

    # --- STEP 10: Sort by Ranking ascending (1 = best) ---
    final_df = final_df.sort_values(by="Ranking", ascending=True, na_position="last")

    # --- STEP 11: Format numeric precision ---
    if "Average Score" in final_df.columns:
        final_df["Average Score"] = final_df["Average Score"].round(2)
    if "Score Range" in final_df.columns:
        final_df["Score Range"] = final_df["Score Range"].round(2)

    # --- STEP 12: Save and format safely ---
    safe_save_excel(final_df.to_excel, output_ranking, index=False)
    format_ranking_table(output_ranking, top_n_to_shade=top_n_to_shade)

    # --- STEP 13: Success message ---
    messagebox.showinfo(
        "Ranking Completed ✅",
        f"The ranking file has been created successfully!"
    )
    print(f"✅ Ranking file saved successfully")


