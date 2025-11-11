
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog, Toplevel, Checkbutton, IntVar, Label, Button, Frame 
from tkinter import font as tkfont
import pandas as pd
from eligibility import create_application_cards
from utils.file_ops import sanitize_filename
import threading
import time
import os, sys
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import eligibility
import scoring
import shutil
import importlib
import json
from documentation import DOC_OVERVIEW, DOC_ELIGIBILITY, DOC_SCORING, DOC_RANKING, DOC_TIPS


def safe_open_excel(path, mode="rb"):
    """
    Try to open an Excel file safely. 
    If the file is already open in Excel, show a clear message instead of PermissionError.
    """
    try:
        return open(path, mode)
    except PermissionError:
        raise PermissionError(
            f"❌ Cannot open '{os.path.basename(path)}'.\n\n"
            f"Please close the file if it's already open in Excel, then try again."
        )
def safe_save_excel(func, *args, **kwargs):
    """
    Wrap Excel save operations to catch PermissionError and show a friendly message.
    Example:
        safe_save_excel(df.to_excel, "output.xlsx")
    """
    try:
        func(*args, **kwargs)
    except PermissionError:
        raise PermissionError(
            "❌ The file could not be saved.\n"
            "Please close it if it's open in Excel, then try again."
        )


def ensure_external_logic_exists():
    """Copy the bundled default logic_config.json to the exe folder if a user-editable copy doesn't exist."""
    import eligibility
    try:
        ext_path = eligibility.get_external_logic_path()
        if not os.path.exists(ext_path):
            # bundled (read-only inside _MEIPASS)
            bundled = eligibility.resource_path("logic_config.json")
            if os.path.exists(bundled):
                shutil.copy(bundled, ext_path)
    except Exception as e:
        # non-fatal; app can still run if developer ensures config presence
        print("Warning ensuring logic file:", e)

def get_font(size=10, weight="normal", slant="roman"):
    """Return a font tuple using the global APP_FONT_NAME."""
    return (APP_FONT_NAME, size, weight)

def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller."""
    try:
        # When the app is run from a PyInstaller bundle
        base_path = sys._MEIPASS
    except Exception:
        # When run from source code
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class ToolTip:
    """Create a tooltip for a given widget."""
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event=None):
        if self.tip_window or not self.text:
            return
        x = self.widget.winfo_rootx() + 25
        y = self.widget.winfo_rooty() + 20
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            tw,
            text=self.text,
            justify="left",
            background="#ffffe0",
            relief="solid",
            borderwidth=1,
            font=("Arial", 10),
            wraplength=350
        )
        label.pack(ipadx=5, ipady=3)

    def hide_tip(self, event=None):
        tw = self.tip_window
        if tw:
            tw.destroy()
        self.tip_window = None



class EdgeFundApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Edge Fund Processor")
        self.root.geometry("850x600")
        self.root.resizable(True, True)

        # ------------------------------
        # Menu Bar
        # ------------------------------
        menubar = tk.Menu(self.root)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="🏠 Home Page", command=self.build_home_page)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        menubar.add_cascade(label="File", menu=file_menu)


        eligibility_menu = tk.Menu(menubar, tearoff=0)
        eligibility_menu.add_command(label="📚 Application Cards & Marking Sheet", command=self.show_eligibility_combined)
        eligibility_menu.add_command(label="⚖️ Final Decision Eligibility", command=self.show_final_decision_content)
        menubar.add_cascade(label="Eligibility", menu=eligibility_menu)

        scoring_menu = tk.Menu(menubar, tearoff=0)
        scoring_menu.add_command(label="📝 Simple Scorecards", command=self.show_scorecards_content)
        scoring_menu.add_command(label="🎯 Scorecards by Topic", command=self.show_scorecards_by_topic_content)
        menubar.add_cascade(label="Scoring", menu=scoring_menu)


        ranking_menu = tk.Menu(menubar, tearoff=0)
        ranking_menu.add_command(label="🏆 Rank Applications", command=self.show_ranking_content)
        menubar.add_cascade(label="Ranking", menu=ranking_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="📖 Instructions", command=self.show_documentation)
        menubar.add_cascade(label="Help", menu=help_menu)


        self.root.config(menu=menubar)

        # ------------------------------
        # State Variables
        # ------------------------------
        self.filepath = None
        self.output_folder = None
        self.columns = []
        self.group_name_column = None
        self.admin_folder = None
        
        self.responses_path = None
        self.name_column_selected = None
        self.selected_columns = {}

        # ------------------------------
        # Main Frame
        # ------------------------------
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(fill="both", expand=True)

        # Home page
        self.build_home_page()

        # Eligibility / Marking Sheet
        self.eligibility_content = None
        self.marking_sheet_content = None

    # =======================================
    # HOME PAGE
    # =======================================
    def build_home_page(self):
        for widget in self.main_frame.winfo_children():
            widget.destroy()
    
        bg_color = COLORS["bg"]
    
        # --- Create scrollable container ---
        container = ttk.Frame(self.main_frame)
        container.pack(fill="both", expand=True)
    
        canvas = tk.Canvas(container, bg=bg_color, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas, style="TFrame")
    
        scrollable_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
    
        canvas.create_window((0, 0), window=scrollable_frame, anchor="n", width=820)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        self._unbind_mousewheel()       # remove any previous binding
        self._bind_mousewheel(canvas)   # bind new one safely

    
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
        # ==================================================
        # 🟦 CONTENT – Logo, Text, Button
        # ==================================================
        logo_path = resource_path(os.path.join("assets", "logo.png"))
        if os.path.exists(logo_path):
            image = Image.open(logo_path).resize((180, 180), Image.LANCZOS)
            logo = ImageTk.PhotoImage(image)
            tk.Label(scrollable_frame, image=logo, bg=bg_color).pack(pady=(20, 15))
            scrollable_frame.image = logo
        else:
            tk.Label(
                scrollable_frame,
                text="Edgy",
                font=get_font(size=16, weight="bold"),
                bg=bg_color
            ).pack(pady=(10, 20))
    
        # Title
        tk.Label(
            scrollable_frame,
            text="Welcome to Edgy – your Edge Fund Application Toolkit",
            font=get_font(size=13, weight="bold"),
            wraplength=700,
            justify="center",
            bg=bg_color
        ).pack(pady=(5, 10))
    
        # Description
        tk.Label(
            scrollable_frame,
            text=(
                "Use the menu bar above or click below to get started.\n\n"
                "Instructions and full documentation are available under Help → 📖 Instructions."
            ),
            font=get_font(size=11),
            wraplength=700,
            justify="center",
            fg="gray",
            bg=bg_color
        ).pack(pady=(0, 15))
    
        # Get Started button
        ttk.Button(
            scrollable_frame,
            text="Get Started →",
            command=self.show_eligibility_combined,
            style="Accent.TButton"
        ).pack(pady=(5, 20))
    
        # ==================================================
        # 📘 FOOTER
        # ==================================================
        fun_sep = tk.Canvas(scrollable_frame, height=2, bg=bg_color, highlightthickness=0)
        fun_sep.pack(fill="x", padx=80, pady=(15, 10))
        fun_sep.create_line(0, 1, 700, 1, fill="#89bdbd", dash=(4, 2), width=2)

    
        footer_text = (
            "Developed with love by Chiara Aquino\n"
            "© 2025 Edge Fund – All rights reserved."
        )
        

        tk.Label(
            scrollable_frame,
            text=footer_text,
            font=get_font(size=9, slant="italic"),
            justify="center",
            fg="gray",
            bg=bg_color
        ).pack(pady=(30, 10))
    
        # 🔧 Ensure layout ready
        self.root.update_idletasks()
    
        # --- Optional: enable scroll with mouse wheel ---
        def _on_mousewheel(event):
            canvas.yview_scroll(-1 * int(event.delta / 120), "units")
    
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        

    # =======================================
    # HELP / DOCUMENTATION WINDOW
    # =======================================
    def show_documentation(self):
        """Open a tabbed, scrollable documentation window (scrolls independently)."""
        doc_win = tk.Toplevel(self.root)
        doc_win.title("Edgy – Documentation")
        doc_win.geometry("900x600")
    
        notebook = ttk.Notebook(doc_win)
        notebook.pack(fill="both", expand=True)
    
        def add_tab(title, text):
            frame = ttk.Frame(notebook)
            notebook.add(frame, text=title)
    
            text_widget = tk.Text(frame, wrap="word", font=get_font(size=10))
            text_widget.insert("1.0", text)
            text_widget.config(state="disabled")
    
            scrollbar = ttk.Scrollbar(frame, orient="vertical", command=text_widget.yview)
            text_widget.config(yscrollcommand=scrollbar.set)
    
            text_widget.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
    
            # --- Independent mousewheel scrolling ---
            def _on_mousewheel(event):
                text_widget.yview_scroll(int(-1 * (event.delta / 120)), "units")
                return "break"  # prevent event from propagating to main window
    
            text_widget.bind("<Enter>", lambda e: text_widget.bind_all("<MouseWheel>", _on_mousewheel))
            text_widget.bind("<Leave>", lambda e: text_widget.unbind_all("<MouseWheel>"))
    
        # --- Create tabs ---
        add_tab("Overview", DOC_OVERVIEW)
        add_tab("Eligibility", DOC_ELIGIBILITY)
        add_tab("Scoring", DOC_SCORING)
        add_tab("Ranking & Outputs", DOC_RANKING)
        add_tab("Tips", DOC_TIPS)


    # =======================================
    # ELIGIBILITY TAB
    # =======================================
    def show_eligibility_combined(self, event=None):
        """Unified interface to create Application Cards and the Eligibility Marking Sheet."""
        for widget in self.main_frame.winfo_children():
            widget.destroy()
    
        # === Scrollable container ===
        container = ttk.Frame(self.main_frame)
        container.pack(fill="both", expand=True)
    
        canvas = tk.Canvas(container)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="n", width=820)
        canvas.configure(yscrollcommand=scrollbar.set)
    
        self._unbind_mousewheel()       # remove any previous scroll binding
        self._bind_mousewheel(canvas)   # bind the new one safely        
    
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
        content = ttk.Frame(scrollable_frame)
        content.pack(anchor="center", pady=20)
    
        # === Title ===
        ttk.Label(
            content,
            text="📚 Create Application Cards and Eligibility Marking Sheet",
            font=get_font(size=14, weight="bold")
        ).pack(pady=10)
    
        # === Step 1: Choose output folder ===
        ttk.Button(
            content, text="Choose folder to save your files",
            command=self.select_output_folder, style="App.TButton"
        ).pack(pady=5)
        self.output_label = tk.Label(content, text="No folder selected", wraplength=600,
                                     fg="gray", font=get_font(size=11))
        self.output_label.pack(pady=5)
    
        # === Step 2: Load responses file ===
        frame_load = ttk.Frame(content)
        frame_load.pack(pady=5, anchor="center")
        
        # Main button
        load_button = ttk.Button(
            frame_load,
            text="Import the file with all the applications",
            command=self.select_file,
            style="App.TButton"
        )
        load_button.pack()
        
        # Frame for the label + info icon (centered)
        file_info_frame = ttk.Frame(content)
        file_info_frame.pack(pady=5, anchor="center")
        
        # Label: "No file selected"
        self.file_label = tk.Label(
            file_info_frame,
            text="No file selected",
            wraplength=600,
            fg="gray",
            font=get_font(size=11)
        )
        self.file_label.pack(side="left", padx=(0, 5))
        
        # ℹ️ Info icon beside the label
        info_icon = ttk.Label(file_info_frame, text="ℹ️", font=("Arial", 12))
        info_icon.pack(side="left")
        
        # Tooltip on hover
        ToolTip(
            info_icon,
            "This should be an Excel file (e.g., ending in .xlsx) that contains all your application responses.\n"
            "Typically, this is the export from Google Forms or another form tool."
        )


    
        # === Step 3: Number of applications ===
        ttk.Label(content, text="Select the maximum number of applications to process:",
                  font=get_font(size=11, weight="bold")).pack(pady=(15, 0))
        self.num_entries_var = tk.StringVar(value="All")
        ttk.Entry(content, textvariable=self.num_entries_var, width=10, font=get_font(size=11)).pack(pady=5)
        ttk.Label(content, text="Type a number (e.g. 20) or 'All'", font=get_font(size=10)).pack(pady=(0, 10))
    
        # === Step 4: Column dropdowns ===
        ttk.Label(content, text="Select the question showing the Organisation Name:",
                  font=get_font(size=11, weight="bold")).pack(pady=(10, 2))
        self.name_column_var = tk.StringVar()
        self.name_column_dropdown = ttk.Combobox(content, textvariable=self.name_column_var,
                                                 state="disabled", width=60)
        self.name_column_dropdown.pack(pady=5)
    
        # --- Application Topics selector (with info icon) ---
        topics_frame = ttk.Frame(content)
        topics_frame.pack(pady=(10, 2))
        
        ttk.Label(
            topics_frame,
            text="Select the question showing the Application Topics:",
            font=get_font(size=11, weight="bold")
        ).pack(side="left", padx=(0, 8))
        
        # ℹ️ Info icon with tooltip
        info_icon_topics = ttk.Label(topics_frame, text="ℹ️", font=("Arial", 12))
        info_icon_topics.pack(side="left")
        
        ToolTip(
            info_icon_topics,
            "These are the categories or topics your applicants selected in their responses.\n\n"
            "If your form doesn’t include topics, tick 'No topics question / skip topics' below."
        )
        
        # Topics dropdown below
        self.topic_column_var = tk.StringVar()
        self.topic_column_dropdown = ttk.Combobox(
            content,
            textvariable=self.topic_column_var,
            state="disabled",
            width=60
        )
        self.topic_column_dropdown.pack(pady=5)
    
        # === Step 5: Option to skip topics ===
        self.no_topics_var = tk.BooleanVar(value=False)
    
        def _toggle_topics_dropdown():
            state = "disabled" if self.no_topics_var.get() else "readonly"
            self.topic_column_dropdown.config(state=state)
    
        ttk.Checkbutton(
            content,
            text="No topics question / skip topics",
            variable=self.no_topics_var,
            command=_toggle_topics_dropdown
        ).pack(pady=(0, 10))
    
    
        # === Step 7: Marking Sheet Configuration Header ===
        ttk.Label(content, text="Build your Eligibility Marking Sheet",
                  font=get_font(size=12, weight="bold"), foreground="#003366").pack(pady=(15, 5))
    
        # === Step 8: Number of scorers ===
        ttk.Label(content, text="Select number of scorers per application:",
                  font=get_font(size=11, weight="bold")).pack(pady=(5, 2))
        self.num_scorers_var = tk.StringVar(value="3")
        num_scorers_dropdown = ttk.Combobox(
            content, textvariable=self.num_scorers_var,
            values=[str(i) for i in range(1, 4)], state="readonly", width=5
        )
        num_scorers_dropdown.pack(pady=(0, 10))
    
        # === Step 9: Unsure / No options using dual selectors ===
        default_unsure = [
            "Some charitable activities", "Focus on mutual aid/ support", "Lack of info online",
            "Some commercial aims", "Mainstream funding", "Works with some institutions",
            "Unsure if led by affected communities", "Other"
        ]
        default_no = [
            "Charity model/activities", "Several paid staff", "Not based in the UK/Ireland",
            "Not aimed towards the UK/Ireland", "Sports Club", "Less than 6 months",
            "Business/ commercial", "More than 20k annual budget", "Entertainment / social only",
            "Less than 3 group members", "Not led by affected community", "Other"
        ]
    
        ttk.Label(content, text="Options for the question: 'If Unsure, tell us why:'",
                  font=get_font(size=11, weight="bold")).pack(pady=(15, 5))
        self.unsure_lists = self.build_dual_option_selector(content, default_unsure)
    
        ttk.Label(content, text="Options for the question: 'If No, tell us why:'",
                  font=get_font(size=11, weight="bold")).pack(pady=(15, 5))
        self.no_lists = self.build_dual_option_selector(content, default_no)
        
        # === Step 6: Select questions for scorer cards ===
        tk.Label(content, text="Select the questions you want scorers to see in their application cards:",
                 font=("Helvetica", 11, "bold")).pack(pady=(10, 0))
    
        self.column_listbox = tk.Listbox(content, selectmode=tk.MULTIPLE, width=80, height=10)
        self.column_listbox.pack(pady=(5, 10))
    
        ttk.Button(content, text="Select All Questions",
                   command=self.select_all_columns).pack(pady=5)
    
        # === Step 10: Generate button ===
        ttk.Button(
            content, text="Create Application Cards and Marking Sheet",
            command=self.run_eligibility_and_marking,
            style="Accent.TButton"
        ).pack(pady=20)
    
        # Navigation buttons
        btn_frame = ttk.Frame(content)
        btn_frame.pack(fill="x", pady=10)
        ttk.Button(btn_frame, text="← Back to Home", command=self.build_home_page, style="Nav.TButton").pack(side="left")
        ttk.Button(btn_frame, text="Go to Final Decision Eligibility →",
                   command=self.show_final_decision_content, style="Nav.TButton").pack(side="right")
    
        self.eligibility_content = content

  

    def create_marking_excel_from_summary(self, group_data, save_path):
        """Generate formatted Eligibility Marking Sheet with dropdowns and colored headers only."""
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.worksheet.datavalidation import DataValidation
    
        num_scorers = int(self.num_scorers_var.get())
        has_topics = any(t for _, _, t in group_data)
    
        # ----------------------------
        # 1️⃣ Build DataFrame
        # ----------------------------
        headers = ["Unique Application ID", "1. Name of your group/organisation"]
        if has_topics:
            headers.append("Topic")
    
        for i in range(1, num_scorers + 1):
            headers += [
                f"Name of scorer {i}",
                "Yes/ No/ Unsure (select from list)",
                "If Unsure, tell us the reason why (select from list)",
                "If No, tell us the reason why (select from list)",
                "Feedback",
            ]
    
        rows = []
        for uid, name, topic in group_data:
            row = [str(uid).zfill(3), name]
            if has_topics:
                row.append(topic)
            row += [""] * (len(headers) - len(row))
            rows.append(row)
    
        df = pd.DataFrame(rows, columns=headers)
    
        # ----------------------------
        # 2️⃣ Write DataFrame with formatting
        # ----------------------------
        writer = pd.ExcelWriter(save_path, engine="xlsxwriter")
        df.to_excel(writer, sheet_name="Eligibility_Marking_Sheet", startrow=1, index=False, header=False)
    
        workbook = writer.book
        worksheet = writer.sheets["Eligibility_Marking_Sheet"]
    
        # === Define formats ===
        base_header_format = {
            "font_name": "Arial", "font_size": 12, "bold": True,
            "align": "center", "valign": "vcenter", "text_wrap": True, "border": 1
        }
    
        # Row fill formats
        yellow_fill = workbook.add_format({"bg_color": "#FFFF99", "border": 1})
        white_fill = workbook.add_format({"bg_color": "#FFFFFF", "border": 1})
    
        # Header formats (colored headers)
        header_yellow = workbook.add_format({**base_header_format, "bg_color": "#FFFF99"})
        header_white = workbook.add_format({**base_header_format, "bg_color": "#FFFFFF"})
        header_blue = workbook.add_format({**base_header_format, "bg_color": "#B7DEE8"})
        header_cream = workbook.add_format({**base_header_format, "bg_color": "#F2F2B2"})
        header_pink = workbook.add_format({**base_header_format, "bg_color": "#E6B8B7"})
    
        # === Apply headers ===
        for col_num, value in enumerate(df.columns.values):
            # Assign header color based on position
            if col_num == 0 or col_num == 1:
                fmt = header_yellow
            elif has_topics and col_num == 2:
                fmt = header_white
            else:
                # Determine color for scorer block headers
                # Each block has 5 columns: blue, cream, pink, pink, pink
                block_pos = (col_num - (3 if has_topics else 2)) % 5
                if block_pos == 0:
                    fmt = header_blue
                elif block_pos == 1:
                    fmt = header_cream
                else:
                    fmt = header_pink
            worksheet.write(0, col_num, value, fmt)
    
        # === Apply row fills ===
        col_count = len(df.columns)
        row_count = len(df)
        # Only color the first two columns' rows
        worksheet.set_column(0, 0, 18, yellow_fill)   # Unique ID
        worksheet.set_column(1, 1, 45, yellow_fill)   # Group Name
        # All other columns white
        for i in range(2, col_count):
            width = 25
            if has_topics and i == 2:
                width = 50
            worksheet.set_column(i, i, width, white_fill)
    
        # Freeze first two columns + header
        worksheet.freeze_panes(1, 2)
    
        writer.close()
    
        # ----------------------------
        # 3️⃣ Add dropdowns (Data Validations)
        # ----------------------------
        wb = load_workbook(save_path)
        ws = wb.active
    
        ws_hidden = wb.create_sheet("_lists")
        ws_hidden.sheet_state = "hidden"
        unsure_options = ["Lack of info", "Conflicting info", "Needs discussion"]
        no_options = ["Not eligible", "Outside scope", "Incomplete application"]
        for i, val in enumerate(unsure_options, start=1):
            ws_hidden.cell(row=i, column=1, value=val)
        for i, val in enumerate(no_options, start=1):
            ws_hidden.cell(row=i, column=2, value=val)
    
        yesno_dv = DataValidation(type="list", formula1='"Yes,No,Unsure"', allow_blank=True)
        unsure_dv = DataValidation(type="list", formula1="=_lists!$A$1:$A$3", allow_blank=True)
        no_dv = DataValidation(type="list", formula1="=_lists!$B$1:$B$3", allow_blank=True)
        ws.add_data_validation(yesno_dv)
        ws.add_data_validation(unsure_dv)
        ws.add_data_validation(no_dv)
    
        fixed_cols = 2 + (1 if has_topics else 0)
        for row in range(2, len(group_data) + 2):
            for scorer_idx in range(num_scorers):
                base_col = fixed_cols + scorer_idx * 5
                yesno_dv.add(ws.cell(row=row, column=base_col + 2))
                unsure_dv.add(ws.cell(row=row, column=base_col + 3))
                no_dv.add(ws.cell(row=row, column=base_col + 4))
    
        wb.save(save_path)

    def run_eligibility_and_marking(self):
        """Generate Application Cards and the Eligibility Marking Sheet in one step."""
        try:
            # --- 1️⃣ Check required inputs ---
            if not hasattr(self, "filepath") or not self.filepath:
                messagebox.showwarning("Missing File", "Please select the Responses file first.")
                return
    
            if not hasattr(self, "output_folder") or not self.output_folder:
                messagebox.showwarning("Missing Folder", "Please choose a folder to save the outputs.")
                return
    
            # --- 2️⃣ Get user selections ---
            org_col = self.name_column_var.get()
            topic_col = self.topic_column_var.get() if not self.no_topics_var.get() else None
            num_scorers = int(self.num_scorers_var.get())
            num_entries = self.num_entries_var.get().strip()
    
            # Determine how many applications to process
            try:
                num_entries = int(num_entries) if num_entries.lower() != "all" else None
            except ValueError:
                messagebox.showerror("Invalid Input", "Please enter a valid number or 'All' for applications.")
                return
    
            # --- 3️⃣ Get selected scorer columns ---
            # --- 4️⃣ Get and remember selected scorer questions ---
            if hasattr(self, "column_listbox"):
                selected_indices = self.column_listbox.curselection()
                if selected_indices:
                    self.selected_scorer_questions = [self.column_listbox.get(i) for i in selected_indices]
            else:
                self.selected_scorer_questions = getattr(self, "selected_scorer_questions", [])
            
            scorer_questions = getattr(self, "selected_scorer_questions", [])
            
            if not scorer_questions:
                confirm = messagebox.askyesno(
                    "No Questions Selected",
                    "No questions were selected for the scorer view.\n\n"
                    "Do you want to continue anyway?"
                )
                if not confirm:
                    return
    
            # --- 4️⃣ Prepare Unsure/No options from dual lists ---
            unsure_options = [self.unsure_lists[1].get(i) for i in range(self.unsure_lists[1].size())]
            no_options = [self.no_lists[1].get(i) for i in range(self.no_lists[1].size())]
    
            # --- 5️⃣ Run generation in background thread with progress bar ---
            from eligibility import generate_application_cards_and_marking
            
            # ✅ Create progress window
            self.progress_window = tk.Toplevel(self.root)
            self.progress_window.title("Generating Application Cards and Marking Sheet...")
            self.progress_window.geometry("420x160")
            tk.Label(
                self.progress_window,
                text="Please wait while the Application Cards and Marking Sheet are being created...",
                wraplength=400
            ).pack(pady=10)
            self.progress_bar = ttk.Progressbar(self.progress_window, length=350, mode="determinate")
            self.progress_bar.pack(pady=10)
            self.progress_label = tk.Label(self.progress_window, text="Starting...")
            self.progress_label.pack()
            
            # ✅ Launch thread
            thread = threading.Thread(
                target=self._generate_eligibility_background,
                args=(org_col, topic_col, num_scorers, num_entries, scorer_questions, unsure_options, no_options),
            )
            thread.start()

    
        except Exception as e:
            messagebox.showerror("Error", f"❌ An unexpected error occurred:\n\n{e}")

    def _update_progress(self, current, total):
        """Update the progress bar percentage."""
        percent = int((current / total) * 100)
        self.root.after(0, lambda: self.progress_label.config(text=f"{percent}% completed"))
        self.root.after(0, lambda: self.progress_bar.config(value=percent))
    
    def _on_generation_complete(self, success=True, error=None):
        """Close progress window and show result."""
        self.root.after(0, self.progress_window.destroy)
        if success:
            messagebox.showinfo("✅ Completed", f"Application Cards and Marking Sheet created!")
        else:
            messagebox.showerror("Error", f"An error occurred during generation:\n\n{error}")
    
    def _generate_eligibility_background(self, org_col, topic_col, num_scorers, num_entries, scorer_questions, unsure_options, no_options):
        """Background thread for generation."""
        try:
            from eligibility import generate_application_cards_and_marking
            generate_application_cards_and_marking(
                responses_path=self.filepath,
                output_folder=self.output_folder,
                group_col=org_col,
                columns_to_select=scorer_questions,
                num_apps=num_entries,
                topics_col=topic_col,
                num_scorers=num_scorers,
                unsure_options=unsure_options,
                no_options=no_options,
                progress_callback=self._update_progress
            )
            self._on_generation_complete(success=True)
        except Exception as e:
            self._on_generation_complete(success=False, error=e)

 

   
    def show_scorecards_content(self):
        """Scoring > Create Simple Scorecards page."""
        self.scorers_file = None
        self.scorers_df = None
        for widget in self.main_frame.winfo_children():
            widget.destroy()
    
        container = ttk.Frame(self.main_frame)
        container.pack(fill="both", expand=True)
    
        canvas = tk.Canvas(container)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
    
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="n", width=820)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        self._unbind_mousewheel()       # remove any previous scroll binding
        self._bind_mousewheel(canvas)   # bind the new one safely
    
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        content = ttk.Frame(scrollable_frame)
        content.pack(anchor="center", pady=20)
    
        ttk.Label(
            content,
            text="📝 Create Simple Scorecards",
            font=get_font(size=14, weight="bold")
        ).pack(pady=10)
        
        # === Step 1: Select Eligible Applications file ===
        section1 = ttk.Frame(content)
        section1.pack(fill="x", pady=10)
        
        # Create a subframe so the label and ℹ️ icon sit side by side
        frame_eligible = ttk.Frame(section1)
        frame_eligible.pack(anchor="w", pady=(0, 2))
        
        # Main label
        ttk.Label(
            frame_eligible,
            text="1. Select only the Eligible Applications:",
            font=get_font(size=11, weight="bold")
        ).pack(side="left")
        
        # Info icon with yellow square background
        info_icon = ttk.Label(frame_eligible, text=" ℹ️", font=("Arial", 11))
        info_icon.pack(side="left", padx=(6, 0))
        
        # Tooltip that explains what file this is
        ToolTip(
            info_icon,
            "This should be the Excel file generated in the previous step ('Final Decision Eligibility').\n"
            "It contains only the applications that were marked as eligible."
        )
        
        # Button to browse for the file
        ttk.Button(
            section1,
            text="Browse...",
            command=self.select_scorecards_eligible_file,
            style="App.TButton"
        ).pack(anchor="w", pady=5)

        
        # Label showing current file status
        self.scorecards_eligible_label = ttk.Label(section1, text="No file selected", foreground="gray")
        self.scorecards_eligible_label.pack(anchor="w")

    
        
    
        # 2) Select output folder
        section2 = ttk.Frame(content)
        section2.pack(fill="x", pady=10)
        ttk.Label(section2, text="2. Choose a location where to save the Scorecards:", font=get_font(size=11, weight="bold")).pack(anchor="w")
        ttk.Button(section2, text="Select Folder", command=self.select_scorecards_output_folder, style="App.TButton").pack(anchor="w", pady=5)
        self.scorecards_output_label = ttk.Label(section2, text="No folder selected", foreground="gray")
        self.scorecards_output_label.pack(anchor="w")
    
        # 3) Add scorers
        section3 = ttk.Frame(content)
        section3.pack(fill="x", pady=10)
        ttk.Label(section3, text="3. Add scorers:", font=get_font(size=11, weight="bold")).pack(anchor="w", pady=(5, 2))
        
        # --- Import scorers table (optional) ---
        import_box = ttk.LabelFrame(section3, text="Import Scorers Table")
        import_box.pack(fill="x", pady=5, padx=5)
        
        import_top = ttk.Frame(import_box)
        import_top.pack(fill="x", pady=5, padx=5)
        
        # Folder button + label
        ttk.Label(import_top, text="Select a file containing the Scorers’ names:").pack(side="left", padx=(0, 10))
        ttk.Button(import_top, text="📂", width=3, command=self.import_scorers_table).pack(side="left")
        self.scorers_table_label = ttk.Label(import_top, text="No file selected", foreground="gray")
        self.scorers_table_label.pack(side="left", padx=8)

        # Column selection section
        ttk.Label(import_box, text="Which column contains the scorers’ names?").pack(anchor="w", padx=10, pady=(5, 2))
        
        self.scorer_column_var = tk.StringVar()
        self.scorer_column_dropdown = ttk.Combobox(import_box, textvariable=self.scorer_column_var,
                                                   state="disabled", width=40)
        self.scorer_column_dropdown.pack(anchor="w", padx=10, pady=(0, 5))
        
        self.load_scorers_button = ttk.Button(
            import_box,
            text="Load Scorers",
            command=self.load_scorers_from_dropdown,
            state="disabled"
        )
        self.load_scorers_button.pack(anchor="w", padx=10, pady=(0, 5))

        
        # --- Manual add/remove scorers (uses same listbox) ---
        manual_box = ttk.LabelFrame(section3, text="Add or Edit Scorers Manually")
        manual_box.pack(fill="x", pady=10, padx=5)
        
        manual_top = ttk.Frame(manual_box)
        manual_top.pack(fill="x", pady=5, padx=5)
        self.scorer_name_var = tk.StringVar()
        ttk.Entry(manual_top, textvariable=self.scorer_name_var, width=40).pack(side="left", padx=(0, 5))
        ttk.Button(manual_top, text="Add Scorer", command=self.add_scorer_to_list).pack(side="left")
        
        # Shared listbox for all scorers
        scorers_list_frame = ttk.Frame(manual_box)
        scorers_list_frame.pack(fill="x", pady=5)
        self.scorers_listbox = tk.Listbox(scorers_list_frame, height=6, width=50)
        self.scorers_listbox.pack(side="left", padx=(0, 5))
        ttk.Button(scorers_list_frame, text="Remove Selected", command=self.remove_selected_scorer).pack(side="left")

    
        # 4) How many times should each application be marked?
        section4 = ttk.Frame(content)
        section4.pack(fill="x", pady=10)
        ttk.Label(section4, text="4. How many times should each application be marked?", font=get_font(size=11, weight="bold")).pack(anchor="w")
        self.reviews_per_app_var = tk.StringVar(value="2")
        ttk.Combobox(
            section4,
            textvariable=self.reviews_per_app_var,
            values=["1", "2", "3"],
            state="readonly",
            width=5
        ).pack(anchor="w", pady=5)
    
        # 5) Configure scorecard columns
        section5 = ttk.Frame(content)
        section5.pack(fill="x", pady=10)
        ttk.Label(section5, text="5. Add new columns to the Scorecard:", font=get_font(size=11, weight="bold")).pack(anchor="w", pady=(0, 5))
    
        # === Scorer Columns Listbox (with horizontal scrollbar) ===
        cols_frame = ttk.Frame(content)
        cols_frame.pack(fill="x", pady=(10, 5))
        
        # Create horizontal scrollbar
        scroll_x = ttk.Scrollbar(cols_frame, orient="horizontal")
        
        # Create the listbox
        self.scorecard_cols_listbox = tk.Listbox(
            cols_frame,
            height=6,
            width=70,
            xscrollcommand=scroll_x.set
        )
        
        # Link scrollbar → listbox
        scroll_x.config(command=self.scorecard_cols_listbox.xview)
        
        # Pack both
        self.scorecard_cols_listbox.pack(side="top", fill="x", padx=(0, 5))
        scroll_x.pack(side="top", fill="x")
        
        # Default columns (same as before)
        default_scorecard_cols = [
            "Points out of 6",
            "Comments - give feedback to explain your score (note you must give a comment for your score to be valid, min 30 words per app)"
        ]
        for col in default_scorecard_cols:
            self.scorecard_cols_listbox.insert(tk.END, col)


    
        cols_btns = ttk.Frame(cols_frame)
        cols_btns.pack(side="left")
        ttk.Button(cols_btns, text="Remove Selected", command=self.remove_scorecard_column).pack(pady=2)
    
        add_col_frame = ttk.Frame(section5)
        add_col_frame.pack(fill="x", pady=5)
        self.new_scorecard_col_var = tk.StringVar()
        ttk.Entry(add_col_frame, textvariable=self.new_scorecard_col_var, width=50).pack(side="left", padx=(0, 5))
        ttk.Button(add_col_frame, text="Add Column", command=self.add_scorecard_column).pack(side="left")
    
        # 6) Generate button + optional report checkbox
        generate_frame = ttk.Frame(content)
        generate_frame.pack(pady=20)
        
        self.generate_report_var = tk.BooleanVar(value=True)
        
        generate_btn = ttk.Button(
            generate_frame,
            text="Create Scorecards",
            command=self.generate_scorecards,
            style="Accent.TButton"
        )
        generate_btn.pack(side="left", padx=(0, 10))
        
        report_check = ttk.Checkbutton(
            generate_frame,
            text="Create a summary report",
            variable=self.generate_report_var
        )
        report_check.pack(side="left")
    
        # Navigation buttons (bottom row)
        nav_frame = ttk.Frame(content)
        nav_frame.pack(fill="x", pady=(30, 10))
    
        left_frame = ttk.Frame(nav_frame)
        left_frame.pack(side="left", fill="x", expand=True)
        ttk.Button(
            left_frame,
            text="← Back to Final Decision Eligibility",
            command=self.show_final_decision_content,
            style="Nav.TButton"
        ).pack(anchor="w", padx=10)
    
        right_frame = ttk.Frame(nav_frame)
        right_frame.pack(side="right", fill="x", expand=True)
        ttk.Button(
            right_frame,
            text="Back to Home",
            command=self.build_home_page,
            style="Nav.TButton"
        ).pack(anchor="e", padx=10)
        
    
    def clear_content(self):
        """Remove all existing widgets from the main content area."""
        for widget in self.main_frame.winfo_children():
            widget.destroy()
    
        # Recreate a fresh frame for new page content
        self.content_frame = ttk.Frame(self.main_frame)
        self.content_frame.pack(fill="both", expand=True, padx=20, pady=20)

    def show_ranking_content(self):
        """Rankings > View Rankings page."""
    
        # 1. Clear previous widgets
        for widget in self.main_frame.winfo_children():
            widget.destroy()
    
        # 2. Create scrollable container
        container = ttk.Frame(self.main_frame)
        container.pack(fill="both", expand=True)
    
        canvas = tk.Canvas(container)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
    
        scrollable_frame.bind(
            "<Configure>", 
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="n", width=820)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        self._unbind_mousewheel()       # remove any previous scroll binding
        self._bind_mousewheel(canvas)   # bind the new one safely
    
    
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
        # 3. Consistent inner frame
        content = ttk.Frame(scrollable_frame)
        content.pack(anchor="center", pady=20)
    
        # === Title (centered, consistent with Scorecards) ===
        ttk.Label(
            content,
            text="🏆 Rank Applications",
            font=get_font(size=14, weight="bold")
        ).pack(pady=(10))
        
        # === 3️⃣ Section 1 – Load Scorers Folder ===
        section1 = ttk.Frame(content)
        section1.pack(fill="x", pady=10, padx=20)
        # Frame to hold label and info icon side by side
        label_frame = ttk.Frame(section1)
        label_frame.pack(anchor="w")
        
        # Main label
        ttk.Label(
            label_frame,
            text="1. Load the folder containing the marked Scorecards:",
            font=get_font(size=11, weight="bold")
        ).pack(side="left")
        
        # Info icon
        info_icon = ttk.Label(label_frame, text=" ℹ️", font=("Arial", 11))
        info_icon.pack(side="left", padx=(6, 0))
        
        # Tooltip for clear instructions
        ToolTip(
            info_icon,
            "Select only the main (parent) folder — the general folder that contains all the individual scorecard subfolders.\n"
            "Do not open or select the subfolders themselves.\n"
            "If selected correctly, a popup window will appear showing the names of all the scorers."
                )

        ttk.Button(section1, text="Select Scorers Folder",
                   command=self.select_scorers_folder_for_ranking,
                   style="App.TButton").pack(anchor="w", pady=5)
        self.scorers_folder_label_ranking = ttk.Label(section1, text="No folder selected", foreground="gray")
        self.scorers_folder_label_ranking.pack(anchor="w")
        self.scorers_count_label = ttk.Label(section1, text="", foreground="gray")
        self.scorers_count_label.pack(anchor="w", pady=(2, 0))
    
        # === 4️⃣ Section 2 – Load Responses Excel ===
        section2b = ttk.Frame(content)
        section2b.pack(fill="x", pady=10, padx=20)
        
        # Frame to hold the label and info icon side by side
        file_info_frame = ttk.Frame(section2b)
        file_info_frame.pack(anchor="w", pady=(0, 2))
        
        # Main label
        ttk.Label(
            file_info_frame,
            text="1. Load Application Responses:",
            font=get_font(size=11, weight="bold")
        ).pack(side="left")
        
        # ℹ️ Info icon beside the label
        info_icon = ttk.Label(file_info_frame, text=" ℹ️", font=("Arial", 12))
        info_icon.pack(side="left", padx=(6, 0))
        
        # Tooltip on hover
        ToolTip(
            info_icon,
            "This should be the original Excel file (e.g., ending in .xlsx) that contains all your application responses as they were submitted.\n"
            "Usually, this is the export from Google Forms or another form tool."
        )
        
        # Load button
        ttk.Button(
            section2b,
            text="Import the file with all the applications",
            command=self.load_responses_excel_for_ranking,
            style="App.TButton"
        ).pack(anchor="w", pady=5)
        
        # Status label
        self.responses_label = ttk.Label(section2b, text="No file loaded", foreground="gray")
        self.responses_label.pack(anchor="w")

    
        # Frame for column checkboxes
        self.columns_frame = ttk.Frame(section2b)
        self.columns_frame.pack(fill="x", pady=5)
    
        # === 5️⃣ Section 3 – Save Location ===
        section3 = ttk.Frame(content)
        section3.pack(fill="x", pady=10, padx=20)
        ttk.Label(section3, text="3. Choose folder to save the Ranking file:",
                  font=get_font(size=11, weight="bold")).pack(anchor="w")
        ttk.Button(section3, text="Select Save Location",
                   command=self.select_output_file_for_ranking,
                   style="App.TButton").pack(anchor="w", pady=5)
        self.output_file_label_ranking = ttk.Label(section3, text="No file selected", foreground="gray")
        self.output_file_label_ranking.pack(anchor="w")
    
        # === 6️⃣ Section 4 – Top N Applications ===
        section4 = ttk.Frame(content)
        section4.pack(fill="x", pady=10, padx=20)
        ttk.Label(section4, text="4. How many applications should pass this round?",
                  font=get_font(size=11, weight="bold")).pack(anchor="w")
        self.num_pass_var = tk.StringVar(value="40")
        ttk.Entry(section4, textvariable=self.num_pass_var, width=10).pack(anchor="w", pady=5)
        ttk.Label(section4,
                  text="(The number of applications you select will be coloured in light green in the ranking table.)",
                  foreground="gray").pack(anchor="w")
    
        # === 7️⃣ Section 5 – Run Ranking ===
        section5 = ttk.Frame(content)
        section5.pack(fill="x", pady=20, padx=20)
        ttk.Button(section5, text="🚀 Generate Ranking File",
                   command=self.run_ranking_process,
                   style="Accent.TButton").pack(anchor="center", pady=5)

    
    
    def select_application_summary_file(self):
        """Select the Application Summary file that will be used to build the Marking Sheet."""
        filepath = filedialog.askopenfilename(
            title="Select the Application Summary file",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if filepath:
            self.summary_path = filepath
            self.summary_label.config(text=os.path.basename(filepath), foreground="black")
    
    def select_marking_save_location(self):
        """Ask where to save the new Eligibility Marking Sheet and pre-fill the filename."""
        default_dir = os.path.dirname(getattr(self, "summary_path", "")) or os.getcwd()
        save_path = filedialog.asksaveasfilename(
            title="Save Eligibility Marking Sheet As",
            initialdir=default_dir,
            initialfile="Eligibility Marking Sheet.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if save_path:
            self.marking_save_path = save_path
            self.marking_save_label.config(text=save_path, foreground="black")
    
    def select_scorecards_eligible_file(self):
        filepath = filedialog.askopenfilename(
            title="Select Eligible Applications file",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not filepath:
            return
    
        self.scorecards_eligible_file = filepath
        if hasattr(self, "scorecards_eligible_label") and self.scorecards_eligible_label.winfo_exists():
            self.scorecards_eligible_label.config(text=os.path.basename(filepath), foreground="black")
        else:
            print("⚠️ Tried to update eligible file label, but it no longer exists (view was changed).")
        
            
        # --- SAFE LOAD of column names ---
        try:
            with safe_open_excel(filepath, "rb") as f:
                df_elig = pd.read_excel(f, nrows=0)
            cols_elig = list(df_elig.columns)
            if hasattr(self, "eligible_group_dropdown"):
                self.eligible_group_dropdown["values"] = cols_elig
                self.eligible_group_dropdown.config(state="readonly")
        except PermissionError as e:
            messagebox.showerror("File in Use", str(e))
            return
        except Exception as e:
            messagebox.showwarning("Warning", f"Could not read eligible file columns:\n{e}")

    def select_scorecards_output_folder(self):
        folder = filedialog.askdirectory(title="Select Output Folder for Scorecards")
        if folder:
            self.scorecards_output_folder = folder
            self.scorecards_output_label.config(text=folder, foreground="black")
    
    
    def add_scorecard_column(self):
        col = self.new_scorecard_col_var.get().strip()
        if col and col not in self.scorecard_cols_listbox.get(0, tk.END):
            self.scorecard_cols_listbox.insert(tk.END, col)
        self.new_scorecard_col_var.set("")
    
    def remove_scorecard_column(self):
        selected = list(self.scorecard_cols_listbox.curselection())
        for i in reversed(selected):
            self.scorecard_cols_listbox.delete(i)
    
    def generate_scorecards(self):
        """Generate matrix and scorecards with a progress bar."""
        try:
            # Check inputs
            if not hasattr(self, "scorecards_eligible_file"):
                messagebox.showwarning("Missing file", "Please select the Eligible Applications file.")
                return
            if not hasattr(self, "scorecards_output_folder"):
                messagebox.showwarning("Missing folder", "Please select an output folder for the Matrix and Scorecards.")
                return
    
            scorers = [self.scorers_listbox.get(i) for i in range(self.scorers_listbox.size())]
            if not scorers:
                messagebox.showwarning("No scorers", "Please add at least one scorer.")
                return
    
            reviews_per_app = int(self.reviews_per_app_var.get())
            scorecard_columns = [self.scorecard_cols_listbox.get(i) for i in range(self.scorecard_cols_listbox.size())]
    
            # Create a popup progress window
            self.progress_window = tk.Toplevel(self.root)
            self.progress_window.title("Generating Scorecards...")
            self.progress_window.geometry("400x150")
            ttk.Label(self.progress_window, text="Creating scorecards... Please wait.").pack(pady=10)
            self.progress = ttk.Progressbar(self.progress_window, mode="determinate", length=300)
            self.progress.pack(pady=10)
            self.progress_label = ttk.Label(self.progress_window, text="0% completed")
            self.progress_label.pack()
            self.progress["value"] = 0
            self.progress["maximum"] = 100
    
            # Run generation in a separate thread
            thread = threading.Thread(
                target=self._generate_scorecards_background,
                args=(scorers, reviews_per_app, scorecard_columns)
            )
            thread.start()
    
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{e}")
            
    def _generate_scorecards_background(self, scorers, reviews_per_app, scorecard_columns):
        """Run scorecard generation in background thread."""
        try:
            total_apps = len(pd.read_excel(self.scorecards_eligible_file))
    
            # Define a progress update callback (optional)
            def progress_callback(current, total):
                percent = int((current / total) * 100)
                self.root.after(0, lambda: self.progress_label.config(text=f"{percent}% completed"))
                self.root.after(0, lambda: self.progress.config(value=percent))
    
            # Try calling scoring with progress_callback if supported
            try:
                scoring.make_matrix_and_scorecards(
                    applications_file=self.scorecards_eligible_file,
                    scorers=scorers,
                    output_folder=self.scorecards_output_folder,
                    reviews_per_app=reviews_per_app,
                    scorecard_columns=scorecard_columns,
                    progress_callback=progress_callback  # Optional arg if implemented
                )
            except TypeError:
                # Fallback if scoring.make_matrix_and_scorecards doesn’t accept callback
                for i in range(0, 101, 5):
                    time.sleep(0.1)
                    self.root.after(0, lambda v=i: self.progress.config(value=v))
                    self.root.after(0, lambda v=i: self.progress_label.config(text=f"{v}% completed"))
                scoring.make_matrix_and_scorecards(
                    applications_file=self.scorecards_eligible_file,
                    scorers=scorers,
                    output_folder=self.scorecards_output_folder,
                    reviews_per_app=reviews_per_app,
                    scorecard_columns=scorecard_columns
                )
    
            # On success
            self.root.after(0, self._on_scorecards_complete, total_apps, len(scorers))
    
        except Exception as err:
            # Capture err safely inside the lambda
            err_msg = str(err)
            def show_error():
                messagebox.showerror("Error", f"An error occurred while creating scorecards:\n{err_msg}")
                self.progress_window.destroy()
            self.root.after(0, show_error)

            
        
    def _on_scorecards_complete(self, total_apps, num_scorers):
        def close_window():
            self.progress_window.destroy()
    
        self.root.after(0, close_window)
        summary_msg = (
            f"✅ Matrix and scorecards created successfully!\n\n"
            f"📦 Total applications: {total_apps}\n"
            f"🗂️  {num_scorers} scorer folders created."
        )
        self.root.after(0, lambda: messagebox.showinfo("Success", summary_msg))

    
    def add_scorer_to_list(self, name=None):
        """Add a scorer to the shared listbox."""
        if name is None:
            name = self.scorer_name_var.get().strip()
        if not name:
            return
        existing = self.scorers_listbox.get(0, tk.END)
        if name not in existing:
            self.scorers_listbox.insert(tk.END, name)
        self.scorer_name_var.set("")
    
    
    def remove_selected_scorer(self):
        """Remove selected scorers from the list."""
        for i in reversed(self.scorers_listbox.curselection()):
            self.scorers_listbox.delete(i)
               
    # -------------------------------
    # SCORECARDS BY TOPIC 
    # -------------------------------
    def show_scorecards_by_topic_content(self, content=None):
        """Scoring > Create Simple Scorecards page."""
        self.scorers_file = None
        self.scorers_df = None
        for widget in self.main_frame.winfo_children():
            widget.destroy()
    
        container = ttk.Frame(self.main_frame)
        container.pack(fill="both", expand=True)
    
        canvas = tk.Canvas(container)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
    
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="n", width=820)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        self._unbind_mousewheel()       # remove any previous scroll binding
        self._bind_mousewheel(canvas)   # bind the new one safely
    
    
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
        content = ttk.Frame(scrollable_frame, padding=20)
        content.pack(anchor="center", pady=20)

    
        # === Title ===
        ttk.Label(content, text="🎯 Create Scorecards by Topic", font=get_font(size=14, weight="bold")).pack(pady=10)
    
        
        # === 1) Select Eligible Applications file ===
        section1 = ttk.Frame(content)
        section1.pack(fill="x", pady=10, padx=20)
        
        # Frame to hold the label and info icon side by side
        frame_eligible = ttk.Frame(section1)
        frame_eligible.pack(anchor="w", pady=(0, 2))
        
        # Main label
        ttk.Label(
            frame_eligible,
            text="1. Select the Eligible Applications file (with Topics):",
            font=get_font(size=11, weight="bold")
        ).pack(side="left")
        
        # Info icon with tooltip
        info_icon = ttk.Label(frame_eligible, text=" ℹ️", font=("Arial", 11))
        info_icon.pack(side="left", padx=(6, 0))
        
        # Tooltip to explain what this file should be
        ToolTip(
            info_icon,
            "This should be the Excel file generated in the previous step ('Final Decision Eligibility').\n"
            "It contains only the applications that were marked as eligible, with a column describing the topics."
        )
        
        # Browse button
        ttk.Button(
            section1,
            text="Browse...",
            command=self.select_eligible_file_for_topic,
            style="App.TButton"
        ).pack(anchor="w", pady=5)
        
        # Label showing file status
        self.eligible_label_topic = ttk.Label(section1, text="No file selected", foreground="gray")
        self.eligible_label_topic.pack(anchor="w")

    
        
        # === 2) Choose output folder ===
        section2 = ttk.Frame(content)
        section2.pack(fill="x", pady=10, padx=20)  # ✅ Already correct
        
        ttk.Label(
            section2,
            text="2. Choose a location where to save the Scorecards:",
            font=get_font(size=11, weight="bold")
        ).pack(anchor="w")
        
        ttk.Button(
            section2,
            text="Select Folder",
            command=self.select_output_folder_for_topic,
            style="App.TButton"
        ).pack(anchor="w", pady=5)
        
        self.output_folder_label_topic = ttk.Label(section2, text="No folder selected", foreground="gray")
        self.output_folder_label_topic.pack(anchor="w")

        # === 2b) Select columns for ID, Group and Topic ===
      
        section2b = ttk.Frame(content)
        section2b.pack(fill="x", pady=10, padx=20)
        ttk.Label(
            section2b,
            text="Select which of the questions that represent Application ID, Organization Name, and Topic(s):",
            font=get_font(size=11, weight="bold")
        ).pack(anchor="w", pady=(0, 5))
        
        cols_frame = ttk.Frame(section2b)
        cols_frame.pack(fill="x", pady=5)
        
        # --- ID column selector ---
        ttk.Label(cols_frame, text="Application ID:").grid(row=0, column=0, sticky="w", padx=5)
        self.id_col_topic_menu = ttk.Combobox(cols_frame, state="readonly", width=30)
        self.id_col_topic_menu.grid(row=0, column=1, padx=5, pady=2)
        
        # --- Group Name column selector ---
        ttk.Label(cols_frame, text="Organization Name:").grid(row=1, column=0, sticky="w", padx=5)
        self.group_col_topic_menu = ttk.Combobox(cols_frame, state="readonly", width=30)
        self.group_col_topic_menu.grid(row=1, column=1, padx=5, pady=2)
        
        # --- Topic column selector ---
        ttk.Label(cols_frame, text="Topics:").grid(row=2, column=0, sticky="w", padx=5)
        self.topic_col_topic_menu = ttk.Combobox(cols_frame, state="readonly", width=30)
        self.topic_col_topic_menu.grid(row=2, column=1, padx=5, pady=2)
        
        # === 🟩 NEW: Topic Loading Section ===
        ttk.Separator(section2b, orient="horizontal").pack(fill="x", pady=10)
        topic_section = ttk.LabelFrame(section2b, text="Topics Setup")
        topic_section.pack(fill="x", pady=10, padx=10)
        
        # Frame to hold the button + info icon side by side
        load_topics_frame = ttk.Frame(topic_section)
        load_topics_frame.pack(pady=5)
        
        # Button to load topics from eligible file (initially disabled)
        self.load_topics_btn = ttk.Button(
            load_topics_frame,
            text="🔄 Load Topics from Eligible Application File",
            command=self.process_topics_from_eligible,
            style="App.TButton",
            state="disabled"  # will be enabled once a topic column is selected
        )
        self.load_topics_btn.pack(side="left")
        
        # ℹ️ Info icon to explain why button is disabled
        info_icon = ttk.Label(load_topics_frame, text=" ℹ️", font=("Arial", 11))
        info_icon.pack(side="left", padx=(6, 0))
        ToolTip(
            info_icon,
            "Select the 'Topics' column above to activate this button."
        )
        
        # 🔴 Instruction label above the checkboxes
        ttk.Label(
            topic_section,
            text='Click on the topic that is considered "general topic" (all scorers can mark this one):',
            font=get_font(size=10, weight="normal"),
            foreground="red"
        ).pack(anchor="w", pady=(10, 2))
        
        # Frame that will hold the dynamically loaded checkboxes
        self.general_topics_frame = ttk.Frame(topic_section)
        self.general_topics_frame.pack(fill="x", pady=5)
        
        # When a topic column is selected → enable the Load Topics button
        def on_topic_column_selected(event=None):
            if self.topic_col_topic_menu.get():
                self.load_topics_btn.config(state="normal")
        
        self.topic_col_topic_menu.bind("<<ComboboxSelected>>", on_topic_column_selected)



        # === 3) Add Scorers (identical to Simple Scorecards) ===
        section3 = ttk.Frame(content)
        section3.pack(fill="x", pady=10,padx=20)
        ttk.Label(section3, text="3. Add the Scorers:", font=get_font(size=11, weight="bold")).pack(anchor="w", pady=(5, 2))
        # 🔹 Add "Open Scorer & Topics Table" button with info icon
        scorer_file_frame = ttk.Frame(section3)
        scorer_file_frame.pack(anchor="w", pady=(0, 10))
        
        open_scorer_btn = ttk.Button(
            scorer_file_frame,
            text="📂 Open Scorer & Topics Table (Previously Saved)",
            command=self.load_scorer_topic_table,
            style="App.TButton"
        )
        open_scorer_btn.pack(side="left")
        
        # Info icon with tooltip
        info_icon = ttk.Label(scorer_file_frame, text=" ℹ️", font=("Arial", 11))
        info_icon.pack(side="left", padx=(6, 0))
        
        ToolTip(
            info_icon,
            "This is the file (ending in .json) that you saved earlier in Step 4 below.\n"
            "If you don't have this file, you can instead get the scorers using a table "
            "that lists their names, or add them manually in the sections below."
        )

        
        import_box = ttk.LabelFrame(section3, text="Import Scorers Table (optional)")
        import_box.pack(fill="x", pady=5, padx=5)
        import_top = ttk.Frame(import_box)
        import_top.pack(fill="x", pady=5, padx=5)
        ttk.Label(import_top, text="Select a file containing the Scorers’ names:").pack(side="left", padx=(0, 10))
        ttk.Button(import_top, text="📂", width=3, command=self.import_scorers_table).pack(side="left")
        self.scorers_table_label = ttk.Label(import_top, text="No file selected", foreground="gray")
        self.scorers_table_label.pack(side="left", padx=8)
    
        ttk.Label(import_box, text="Which column contains the scorers’ names?").pack(anchor="w", padx=10, pady=(5, 2))
        self.scorer_column_var = tk.StringVar()
        self.scorer_column_dropdown = ttk.Combobox(import_box, textvariable=self.scorer_column_var,
                                                   state="disabled", width=40)
        self.scorer_column_dropdown.pack(anchor="w", padx=10, pady=(0, 5))
        self.load_scorers_button = ttk.Button(import_box, text="Load Scorers",
                                              command=self.load_scorers_from_dropdown, state="disabled")
        self.load_scorers_button.pack(anchor="w", padx=10, pady=(0, 5))
    
        manual_box = ttk.LabelFrame(section3, text="Add or Edit Scorers Manually")
        manual_box.pack(fill="x", pady=10, padx=5)
        manual_top = ttk.Frame(manual_box)
        manual_top.pack(fill="x", pady=5, padx=5)
        self.scorer_name_var = tk.StringVar()
        ttk.Entry(manual_top, textvariable=self.scorer_name_var, width=40).pack(side="left", padx=(0, 5))
        ttk.Button(manual_top, text="Add the Scorers", command=self.add_scorer_to_list).pack(side="left")
    
        scorers_list_frame = ttk.Frame(manual_box)
        scorers_list_frame.pack(fill="x", pady=5)
        self.scorers_listbox = tk.Listbox(scorers_list_frame, height=6, width=50)
        self.scorers_listbox.pack(side="left", padx=(0, 5))
        ttk.Button(scorers_list_frame, text="Remove Selected",
                   command=self.remove_selected_scorer).pack(side="left")
    
        # === 4) Assign Topics to Scorers ===
        section4 = ttk.Frame(content)
        section4.pack(fill="x", pady=10, padx=20)
        
        ttk.Label(
            section4,
            text="4. Assign Topics to Scorers:",
            font=get_font(size=11, weight="bold")
        ).pack(anchor="w")
        
        # --- Buttons for scorer-topic management ---
        buttons_frame = ttk.Frame(section4)
        buttons_frame.pack(fill="x", pady=5)
        buttons_frame.grid_columnconfigure(1, weight=1)
        
        ttk.Button(
            buttons_frame,
            text="🔧Create the Scorers-Topics Table",
            command=self.build_scorer_topic_ui
        ).grid(row=0, column=0, sticky="w", padx=(0, 10))
        
        ttk.Button(
            section4,
            text="💾 Save the Scorers-Topics Table",
            command=self.save_scorer_topic_table
        ).pack(anchor="w", pady=5)
        
        # Frame where the scorer–topic table will appear
        self.topic_assign_frame = ttk.Frame(section4)
        self.topic_assign_frame.pack(fill="both", expand=True, pady=5)

    
        # === 5) Reviews per app ===
        section5 = ttk.Frame(content)
        section5.pack(fill="x", pady=10,padx=20)
        ttk.Label(section5, text="5. How many times should each application be marked?",
                  font=get_font(size=11, weight="bold")).pack(anchor="w")
        self.reviews_per_app_var = tk.StringVar(value="2")
        ttk.Combobox(section5, textvariable=self.reviews_per_app_var,
                     values=["1", "2", "3"], state="readonly", width=5).pack(anchor="w", pady=5)
    
        # === 6) Scorecard columns ===
        section6 = ttk.Frame(content)
        section6.pack(fill="x", pady=10, padx=20)
        
        ttk.Label(
            section6,
            text="6. Add new columns to the Scorecard:",
            font=get_font(size=11, weight="bold")
        ).pack(anchor="w", pady=(0, 5))
        
        # === Scorer Columns Listbox (with horizontal scrollbar) ===
        cols_frame = ttk.Frame(content)
        cols_frame.pack(fill="x", pady=(10, 5))
        
        # Create horizontal scrollbar
        scroll_x = ttk.Scrollbar(cols_frame, orient="horizontal")
        
        # Create the listbox
        self.scorecard_cols_listbox = tk.Listbox(
            cols_frame,
            height=6,
            width=70,
            xscrollcommand=scroll_x.set
        )
        
        # Link scrollbar → listbox
        scroll_x.config(command=self.scorecard_cols_listbox.xview)
        
        # Pack both
        self.scorecard_cols_listbox.pack(side="top", fill="x", padx=(0, 5))
        scroll_x.pack(side="top", fill="x")
        
        # Default columns (same as before)
        default_scorecard_cols = [
            "Points out of 6",
            "Comments - give feedback to explain your score (note you must give a comment for your score to be valid, min 30 words per app)"
        ]
        for col in default_scorecard_cols:
            self.scorecard_cols_listbox.insert(tk.END, col)

        
        cols_btns = ttk.Frame(cols_frame)
        cols_btns.pack(side="left")
        ttk.Button(cols_btns, text="Remove Selected", command=self.remove_scorecard_column).pack(pady=2)
        
        # Add Column (bottom row)
        add_col_frame = ttk.Frame(section6)
        add_col_frame.pack(fill="x", pady=5)
        self.new_scorecard_col_var = tk.StringVar()
        ttk.Entry(add_col_frame, textvariable=self.new_scorecard_col_var, width=50).pack(side="left", padx=(0, 5))
        ttk.Button(add_col_frame, text="Add Column", command=self.add_scorecard_column).pack(side="left")

    
        # === 7) Generate button + report option ===
        generate_frame = ttk.Frame(content)
        generate_frame.pack(pady=20)
        self.create_report_var = tk.BooleanVar(value=True)
        ttk.Button(generate_frame, text="Create Scorecards",
                   command=self.run_topic_scorecards, style="Accent.TButton").pack(side="left", padx=(0, 10))
        ttk.Checkbutton(generate_frame, text="Create a summary report",
                        variable=self.create_report_var).pack(side="left")

        
    def select_eligible_file_for_topic(self):
        path = filedialog.askopenfilename(
            title="Select Eligible Applications file",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not path:
            return
        self.scorecards_eligible_file = path
        self.eligible_label_topic.config(text=os.path.basename(path))
        self.load_eligible_columns_for_topic()

    def load_eligible_columns_for_topic(self):
        """Load column names from the Eligible Applications file into the dropdowns."""
        if not hasattr(self, "scorecards_eligible_file"):
            return
        try:
            with safe_open_excel(self.scorecards_eligible_file, "rb") as f:
                df = pd.read_excel(f, nrows=1)
            cols = list(df.columns)
            if not cols:
                messagebox.showerror("Error", "No columns found in the selected Excel file.")
                return
    
            # ✅ Populate all three dropdowns (ID, Group, Topic)
            self.id_col_topic_menu["values"] = cols
            self.group_col_topic_menu["values"] = cols
            self.topic_col_topic_menu["values"] = cols 
    
            # Enable dropdowns
            self.id_col_topic_menu.config(state="readonly")
            self.group_col_topic_menu.config(state="readonly")
            self.topic_col_topic_menu.config(state="readonly")
            
            # Optional defaults
            if len(cols) >= 3:
                self.id_col_topic_menu.current(0)
                self.group_col_topic_menu.current(1)
                self.topic_col_topic_menu.current(2)
            elif len(cols) == 2:
                self.id_col_topic_menu.current(0)
                self.group_col_topic_menu.current(1)
            else:
                self.id_col_topic_menu.current(0)
    
            messagebox.showinfo("Columns Loaded", "✅ Columns loaded successfully from the selected file.")
        except PermissionError as e:
            messagebox.showerror("File in Use", str(e))
            return
        except Exception as e:
            messagebox.showerror("Error loading columns", f"Could not read the Excel file:\n{e}")

    def select_output_folder_for_topic(self):
        from tkinter import filedialog
        folder = filedialog.askdirectory(title="Select Folder")
        if folder:
            self.scorecards_output_folder = folder
            self.output_folder_label_topic.config(text=folder)
    
    def create_scorers_and_topics_ui(self, parent):
        """Build scorer list, topic loading, topic assignment, and settings section."""
        # === Scorers ===
        section = ttk.LabelFrame(parent, text="Scorers Setup")
        section.pack(fill="x", pady=10, padx=10)
    
        ttk.Label(section, text="Add or Import Scorers:", font=get_font(size=11, weight="bold")).pack(anchor="w", pady=5)
    
        # Manual add
        add_frame = ttk.Frame(section)
        add_frame.pack(fill="x", pady=5)
        self.scorer_name_var = tk.StringVar()
        ttk.Entry(add_frame, textvariable=self.scorer_name_var, width=40).pack(side="left", padx=(0, 5))
        ttk.Button(add_frame, text="Add Scorer", command=self.add_scorer_to_list).pack(side="left")
        ttk.Button(add_frame, text="Remove Selected", command=self.remove_selected_scorer).pack(side="left", padx=5)
    
        # Listbox
        self.scorers_listbox = tk.Listbox(section, height=6, width=50)
        self.scorers_listbox.pack(pady=5)
    
        # Import scorers file
        ttk.Button(section, text="📂 Import Scorers Table", command=self.import_scorers_table).pack(pady=5)
    
        # === Topic Loading ===
        ttk.Separator(parent, orient="horizontal").pack(fill="x", pady=10)
        topic_section = ttk.LabelFrame(parent, text="Topics Setup")
        topic_section.pack(fill="x", pady=10, padx=10)
    
        ttk.Button(topic_section, text="🔄 Load Topics from Eligible File", command=self.process_topics_from_eligible).pack(pady=5)
    
        # Frame to hold general topics checkboxes
        self.general_topics_frame = ttk.Frame(topic_section)
        self.general_topics_frame.pack(fill="x", pady=5)
    
        # === Scorer–Topic Assignment ===
        ttk.Button(topic_section, text="🧩 Build Scorer–Topic Assignment Table", command=self.build_scorer_topic_ui).pack(pady=5)
        ttk.Button(topic_section, text="📂 Open Scorer Table (Previously Saved)", command=self.load_scorer_topic_table).pack(anchor="w", pady=5)
        self.topic_assign_frame = ttk.Frame(topic_section)
        self.topic_assign_frame.pack(fill="both", expand=True, pady=5)
        
        ttk.Button(topic_section, text="💾 Save Current Scorer Table", command=self.save_scorer_topic_table).pack(anchor="w", pady=5)

    
        # === Reviews per app ===
        ttk.Separator(parent, orient="horizontal").pack(fill="x", pady=10)
        ttk.Label(parent, text="How many times should each application be marked?",
                  font=get_font(size=11, weight="bold")).pack(anchor="w", padx=10)
        self.reviews_per_app_var = tk.StringVar(value="2")
        ttk.Combobox(parent, textvariable=self.reviews_per_app_var, values=["1", "2", "3"], state="readonly", width=5).pack(anchor="w", padx=10, pady=5)
    
        # === Scorecard columns ===
        ttk.Label(parent, text="Columns to include in scorecards:", font=get_font(size=11, weight="bold")).pack(anchor="w", padx=10, pady=(10, 0))
        self.scorecard_cols_listbox = tk.Listbox(parent, height=5, width=70)
        self.scorecard_cols_listbox.pack(anchor="w", padx=10, pady=5)
        for col in ["Points out of 6", "Comments (min 30 words)"]:
            self.scorecard_cols_listbox.insert(tk.END, col)
    
        col_entry_frame = ttk.Frame(parent)
        col_entry_frame.pack(anchor="w", padx=10, pady=5)
        self.new_scorecard_col_var = tk.StringVar()
        ttk.Entry(col_entry_frame, textvariable=self.new_scorecard_col_var, width=50).pack(side="left", padx=(0, 5))
        ttk.Button(col_entry_frame, text="Add Column", command=self.add_scorecard_column).pack(side="left")
        ttk.Button(col_entry_frame, text="Remove Selected", command=self.remove_scorecard_column).pack(side="left", padx=5)
    
        # === Report Option ===
        self.create_report_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(parent, text="Create a summary report", variable=self.create_report_var).pack(anchor="w", padx=10, pady=10)

    
    
    
    
    def run_topic_scorecards(self):
        from scoring import make_topic_based_scorecards
        if not hasattr(self, "scorecards_output_folder") or not self.scorecards_output_folder:
            messagebox.showwarning("Missing Folder", "Please select an output folder before generating scorecards.")
            return
        
        if not hasattr(self, "scorecards_eligible_file") or not self.scorecards_eligible_file:
            messagebox.showwarning("Missing File", "Please select the Eligible Applications file first.")
            return

        # --- Progress Bar setup (same as Simple Scorecards) ---
        progress_win = tk.Toplevel(self.root)
        progress_win.title("Generating Topic-Based Scorecards")
        progress_win.geometry("400x120")
        progress_win.resizable(False, False)
        
        ttk.Label(progress_win, text="Generating topic-based scorecards...", font=get_font(size=11)).pack(pady=(20, 10))
        progress = ttk.Progressbar(progress_win, mode="indeterminate", length=300)
        progress.pack(pady=5)
        progress.start()
        self.root.update_idletasks()

    
        scorers_topics = self.get_scorers_topics()  # whatever you use to read scorer-topic input
        general_topics = self.get_general_topics()
        scorecard_columns = self.get_scorecard_columns()
    
        def run_task():
            try:
                make_topic_based_scorecards(
                eligible_file=self.scorecards_eligible_file,
                output_folder=self.scorecards_output_folder,
                scorers_topics=self.get_scorers_topics(),
                id_column=self.id_col_topic_menu.get(),
                topic_column=self.topic_col_topic_menu.get(),
                group_col_eligible=self.group_col_topic_menu.get(),
                general_topics=self.get_general_topics(),
                reviews_per_app=int(self.reviews_per_app_var.get()),
                scorecard_columns=self.get_scorecard_columns(),
                create_report=self.create_report_var.get()
                    )

                self.root.after(0, lambda: [
                    progress.stop(),
                    progress_win.destroy(),
                    messagebox.showinfo("Success", "✅ Topic-based scorecards created successfully!")
                ])
            except Exception as e:
                err_msg = str(e)
                self.root.after(0, lambda msg=err_msg: [
                    progress.stop(),
                    progress_win.destroy(),
                    messagebox.showerror("Error", f"An error occurred:\n{msg}")
                ])

        threading.Thread(target=run_task, daemon=True).start()





    def process_topics_from_eligible(self):
        """Extract unique topic names from the selected Topic column in the Eligible Applications file."""
        if not hasattr(self, "scorecards_eligible_file"):
            messagebox.showwarning("Missing file", "Please select the Eligible Applications file first.")
            return
        topic_col = self.topic_col_topic_menu.get()
        if not topic_col:
            messagebox.showwarning("Missing selection", "Please select which column contains topics.")
            return
    
        try:
            df = pd.read_excel(self.scorecards_eligible_file, usecols=[topic_col])
            all_topics = []
            for val in df[topic_col].dropna():
                parts = [p.strip() for p in str(val).split(",") if p.strip()]
                all_topics.extend(parts)
            unique_topics = sorted(set(all_topics))
            self.available_topics = unique_topics
    
            # Clear previous widgets
            for widget in self.general_topics_frame.winfo_children():
                widget.destroy()
    
            ttk.Label(
                self.general_topics_frame,
                text="Select general topics (optional):",
                font=get_font(size=11, weight="bold")
            ).pack(anchor="w", padx=10, pady=5)
    
            self.general_topic_vars = {}
            for t in self.available_topics:
                var = tk.BooleanVar()
                cb = ttk.Checkbutton(self.general_topics_frame, text=t, variable=var)
                cb.pack(anchor="w", padx=20)
                self.general_topic_vars[t] = var
    
            # Add manual topic input
            add_frame = ttk.Frame(self.general_topics_frame)
            add_frame.pack(fill="x", pady=5, padx=15)
            self.new_topic_var = tk.StringVar()
            ttk.Entry(add_frame, textvariable=self.new_topic_var, width=50).pack(side="left", padx=(0, 5))
            ttk.Button(add_frame, text="Add Topic", command=self.add_new_topic).pack(side="left")
    
            messagebox.showinfo("Topics Loaded", f"✅ {len(unique_topics)} unique topics loaded from '{topic_col}'.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read topics:\n{e}")
            
    def add_new_topic(self):
        """Allow user to add a new topic manually."""
        new_topic = self.new_topic_var.get().strip()
        if not new_topic:
            return
        if new_topic not in getattr(self, "available_topics", []):
            self.available_topics.append(new_topic)
            var = tk.BooleanVar()
            cb = ttk.Checkbutton(self.general_topics_frame, text=new_topic, variable=var)
            cb.pack(anchor="w", padx=20)
            self.general_topic_vars[new_topic] = var
        self.new_topic_var.set("")

  
    
    def build_scorer_topic_ui(self):
        """Show one column per scorer, each with topics listed vertically and checkboxes beside labels."""
        # Remove all existing topic assignment frames except the refresh button
        for widget in self.topic_assign_frame.winfo_children():
            if not isinstance(widget, ttk.Button):
                widget.destroy()
    
        scorers = [self.scorers_listbox.get(i) for i in range(self.scorers_listbox.size())]
        if not scorers:
            messagebox.showwarning("No scorers", "Please add or import scorers first.")
            return
        if not self.available_topics:
            messagebox.showwarning("No topics", "Please load topics first.")
            return
    
        ttk.Label(
            self.topic_assign_frame,
            text="Assign topics to each scorer (scroll → if needed)",
            font=get_font(size=11, weight="bold")
        ).pack(anchor="w", padx=10, pady=(5, 5))
    
        # --- Scrollable container for topic assignment ---
        container = ttk.Frame(self.topic_assign_frame)
        container.pack(fill="both", expand=True, padx=10, pady=5)
        
        # Slightly smaller blue area (was height=350 → now 250)
        canvas = tk.Canvas(
            container,
            bg="#E3F2FD",          # light blue background
            highlightthickness=0,
            height=250
        )
        canvas.pack(side="left", fill="both", expand=True)
        
        # Larger, more visible scrollbars
        style = ttk.Style()
        style.configure("TScrollbar", arrowsize=20)  # increase arrow button size
        
        scrollbar_x = ttk.Scrollbar(container, orient="horizontal", command=canvas.xview)
        scrollbar_x.pack(side="bottom", fill="x")
        scrollbar_y = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollbar_y.pack(side="right", fill="y")
        
        canvas.configure(xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set)
        
        # Internal frame (holds each scorer’s topic list)
        inner_frame = ttk.Frame(canvas, padding=5)  # reduced padding
        canvas.create_window((0, 0), window=inner_frame, anchor="nw")
        
        def update_scrollregion(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        inner_frame.bind("<Configure>", update_scrollregion)

    
        # dictionary for checkbox vars
        self.scorer_topic_vars = {scorer: {} for scorer in scorers}
    
        # Build one column per scorer
        for col, scorer in enumerate(scorers):
            frame = ttk.LabelFrame(inner_frame, text=scorer)
            frame.grid(row=0, column=col, padx=10, pady=5, sticky="n")
    
            for topic in self.available_topics:
                row = ttk.Frame(frame)
                row.pack(anchor="w", pady=2, padx=5)
                ttk.Label(row, text=topic, width=30, anchor="w").pack(side="left")
                var = tk.BooleanVar()
                chk = ttk.Checkbutton(row, variable=var)
                chk.pack(side="left", padx=(5, 0))
                self.scorer_topic_vars[scorer][topic] = var
    
        # Allow horizontal scrolling with Shift + mouse wheel
        def _on_shiftwheel(event):
            canvas.xview_scroll(-1 * int(event.delta / 120), "units")
    
        # Vertical scroll normally
        def _on_mousewheel(event):
            canvas.yview_scroll(-1 * int(event.delta / 120), "units")
    
        canvas.bind_all("<Shift-MouseWheel>", _on_shiftwheel)
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

    def save_scorer_topic_table(self):
        """Save the scorer–topic assignment table as JSON."""
        try:
            if not hasattr(self, "scorer_topic_vars"):
                messagebox.showwarning("No data", "Please build the scorer–topic assignment table first.")
                return
            scorers_topics = self.get_scorers_topics()
            default_name = "Scorers Names and Topics.json"
            
            path = filedialog.asksaveasfilename(
                initialfile=default_name,
                defaultextension=".json",
                filetypes=[("JSON Files", "*.json")],
                title="Save Scorer–Topic Table As"
            )
            if not path:
                return
            with open(path, "w", encoding="utf-8") as f:
                json.dump(scorers_topics, f, indent=2)
            messagebox.showinfo("Saved", f"✅ Scorer–Topic table saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save scorer table:\n{e}")
    
    def load_scorer_topic_table(self):
        """Load a previously saved scorer–topic assignment JSON."""
        try:
            path = filedialog.askopenfilename(
                filetypes=[("JSON Files", "*.json")],
                title="Open Previously Saved Scorer–Topic Table"
            )
            if not path:
                return
            with open(path, "r", encoding="utf-8") as f:
                scorers_topics = json.load(f)
    
            # Ensure scorers exist in the listbox
            current_scorers = set(self.scorers_listbox.get(0, tk.END))
            for s in scorers_topics.keys():
                if s not in current_scorers:
                    self.scorers_listbox.insert(tk.END, s)
    
            # Ensure topics are known
            if not hasattr(self, "available_topics") or not self.available_topics:
                all_topics = sorted({t for lst in scorers_topics.values() for t in lst})
                self.available_topics = all_topics
    
            # Rebuild UI checkboxes with loaded data
            self.build_scorer_topic_ui()
            for s, topics in scorers_topics.items():
                for t in topics:
                    if s in self.scorer_topic_vars and t in self.scorer_topic_vars[s]:
                        self.scorer_topic_vars[s][t].set(True)
            messagebox.showinfo("Loaded", f"✅ Scorer–Topic table loaded from:\n{os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load scorer table:\n{e}")


    def generate_scorecards_by_topic(self):
        try:
            # Collect scorer–topic mapping
            scorers_topics = {
                s: [self.scorer_topic_map[s].get(i) for i in self.scorer_topic_map[s].curselection()]
                for s in self.scorer_topic_map
            }
            
            # Check for scorers with no assigned topics
            no_topic_scorers = [s for s, topics in scorers_topics.items() if not topics]
            if no_topic_scorers:
                proceed = messagebox.askyesno(
                    "Unassigned Topics",
                    f"The following scorers have not been assigned any topics:\n\n"
                    + "\n".join(no_topic_scorers)
                    + "\n\nThey will not be assigned any applications.\nProceed anyway?"
                )
                if not proceed:
                    return

            
            
            general_topics = [t for t, var in self.general_topic_vars.items() if var.get()]
            scorecard_columns = [self.scorecard_cols_listbox.get(i) for i in range(self.scorecard_cols_listbox.size())]
    
            # --- Progress window setup ---
            self.progress_window = tk.Toplevel(self.root)
            self.progress_window.title("Generating Topic-Based Scorecards")
            self.progress_window.geometry("420x140")
            ttk.Label(self.progress_window, text="Creating topic-based scorecards, please wait...").pack(pady=10)
            self.progress = ttk.Progressbar(self.progress_window, orient="horizontal", length=300, mode="determinate")
            self.progress.pack(pady=10)
            self.progress_label = ttk.Label(self.progress_window, text="Starting...")
            self.progress_label.pack()
    
            # Run background thread
            threading.Thread(
                target=self._generate_scorecards_by_topic_background,
                args=(scorers_topics, general_topics, scorecard_columns),
                daemon=True
            ).start()
    
        except Exception as e:
            messagebox.showerror("Error", f"Error while preparing topic-based scorecards:\n{e}")

   

    # -------------------------------
    # FINAL DECISION ELIGIBILITY
    # -------------------------------
    def show_final_decision_content(self):
        for widget in self.main_frame.winfo_children():
            widget.destroy()
    
        container = ttk.Frame(self.main_frame)
        container.pack(fill="both", expand=True)
    
        canvas = tk.Canvas(container)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
    
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="center", width=820)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        self._unbind_mousewheel()       # remove any previous scroll binding
        self._bind_mousewheel(canvas)   # bind the new one safely
    
    
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
        content = ttk.Frame(scrollable_frame)
        content.pack(anchor="center", pady=20)
    
        ttk.Label(content, text="⚖️ Final Decision Eligibility", font=get_font(size=14, weight="bold")).pack(pady=10)
    
        # Select input marking sheet
        ttk.Button(content, text="Import the Eligibility Marking Sheet", command=self.select_final_input).pack(pady=5)
        self.final_input_label = ttk.Label(content, text="No file selected", font=get_font(size=11), foreground="gray")
        self.final_input_label.pack(pady=5)
    
        # Select output folder
        ttk.Button(content, text="Choose the location where to save eligibility results", command=self.select_final_output_folder).pack(pady=5)
        self.final_output_label = ttk.Label(content, text="No folder selected", font=get_font(size=11), foreground="gray")
        self.final_output_label.pack(pady=5)
    
        # Change logic button
        ttk.Button(content, text="Change Marking Logic", command=self.show_logic_editor).pack(pady=10)
 
        ttk.Button(content, text="Generate Final Decision", command=self.generate_final_decision, style="Accent.TButton").pack(pady=15)
    
        # === Navigation buttons (bottom row) ===
        nav_frame = ttk.Frame(content)
        nav_frame.pack(fill="x", pady=(30, 10))
        
        left_frame = ttk.Frame(nav_frame)
        left_frame.pack(side="left", fill="x", expand=True)
        ttk.Button(
            left_frame,
            text="← Back to Final Decision Eligibility",
            command=self.show_final_decision_content,
            style="Nav.TButton"
        ).pack(anchor="w", padx=10)
        
        right_frame = ttk.Frame(nav_frame)
        right_frame.pack(side="right", fill="x", expand=True)
        ttk.Button(
            right_frame,
            text="Back to Home",
            command=self.build_home_page,
            style="Nav.TButton"
        ).pack(anchor="e", padx=10)


        self.final_decision_content = content
    
    def select_final_input(self):
        filepath = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if filepath:
            self.final_input_file = filepath
            self.final_input_label.config(text=os.path.basename(filepath), foreground="black")
    
    def select_final_output_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.final_output_folder = folder
            self.final_output_label.config(text=folder, foreground="black")
    
    def show_logic_editor(self):
        """
        Open a window showing all unique keys from logic_config.json and a dropdown for each result.
        Only the right-hand result is editable.
        """
        # Determine path to external config (same helper from eligibility)
        try:
            logic_data = eligibility.load_logic_config()
        except Exception as e:
            messagebox.showerror("Load error", f"Could not load logic config:\n{e}")
            return
    
        self._logic_data_working = logic_data  # keep in memory
    
        logic_window = tk.Toplevel(self.root)
        logic_window.title("Marking Logic Editor")
        logic_window.geometry("700x520")
    
        tk.Label(logic_window, text="Edit the marking logic",
                 font=get_font(size=12, weight="bold")).pack(pady=8)
    
        # ⚠️ Warning about logic changes
        warning_text = (
            "⚠️ Attention: If you change the marking logic, "
            "your changes will be saved for all future sessions.\n"
            "The previous logic will be permanently overwritten."
        )
        
        tk.Label(
            logic_window,
            text=warning_text,
            font=get_font(size=10),
            fg="red",
            justify="center",
            wraplength=700,
            bg=COLORS["bg"]
        ).pack(pady=(0, 10))

        # Generate button
        
        notebook = ttk.Notebook(logic_window)
        notebook.pack(fill="both", expand=True, padx=8, pady=8)
    
        # 3 scorers tab
        frame3 = ttk.Frame(notebook)
        notebook.add(frame3, text="3 Scorers")
        self.logic3_vars = {}
        # Sorted keys for consistent ordering in the UI
        keys3 = sorted(logic_data.get("LOGIC_3SCORERS", {}).keys())
        for k in keys3:
            row = ttk.Frame(frame3)
            row.pack(anchor="w", pady=2, padx=6)
            label_text = k.replace(",", " + ")
            ttk.Label(row, text=f"{label_text}  =", width=30).pack(side="left", padx=4)
            var = tk.StringVar(value=logic_data["LOGIC_3SCORERS"][k])
            cb = ttk.Combobox(row, textvariable=var, values=["yes", "no"], width=10, state="readonly")
            cb.pack(side="left", padx=4)
            self.logic3_vars[k] = var
    
        # 2 scorers tab
        frame2 = ttk.Frame(notebook)
        notebook.add(frame2, text="2 Scorers")
        self.logic2_vars = {}
        keys2 = sorted(logic_data.get("LOGIC_2SCORERS", {}).keys())
        for k in keys2:
            row = ttk.Frame(frame2)
            row.pack(anchor="w", pady=2, padx=6)
            label_text = k.replace(",", " + ")
            ttk.Label(row, text=f"{label_text}  =", width=30).pack(side="left", padx=4)
            var = tk.StringVar(value=logic_data["LOGIC_2SCORERS"][k])
            cb = ttk.Combobox(row, textvariable=var, values=["yes", "no"], width=10, state="readonly")
            cb.pack(side="left", padx=4)
            self.logic2_vars[k] = var
    
        btn_frame = ttk.Frame(logic_window)
        btn_frame.pack(fill="x", pady=10)
        ttk.Button(btn_frame, text="💾 Save Changes", command=lambda: self.save_logic_changes(logic_window)).pack(side="right", padx=8)
        ttk.Button(btn_frame, text="Cancel", command=logic_window.destroy).pack(side="right", padx=8)

    def save_logic_changes(self, window):
        """
        Collect updated results and save them to the external logic_config.json,
        then reload the eligibility module so new logic takes immediate effect.
        """
        updated_logic = {
            "LOGIC_3SCORERS": {k: v.get() for k, v in self.logic3_vars.items()},
            "LOGIC_2SCORERS": {k: v.get() for k, v in self.logic2_vars.items()}
        }
    
        try:
            # Save using eligibility.save_logic_config (writes into exe dir external file)
            eligibility.save_logic_config(updated_logic)
        except Exception as e:
            messagebox.showerror("Save error", f"Failed to save logic config:\n{e}")
            return
    
        # Try to reload eligibility module so changes apply immediately
        try:
            importlib.reload(eligibility)
        except Exception as e:
            # Not fatal — saved to disk; user may restart
            messagebox.showwarning("Reload warning", f"Saved but failed to reload module:\n{e}\nYou may need to restart the app.")
            window.destroy()
            return
    
        window.destroy()
        messagebox.showinfo("Saved", "✅ Logic updated and reloaded successfully!")

    def get_scorers_topics(self):
        """
        Read all topic assignments from self.scorer_topic_vars and return
        a dictionary mapping each scorer -> list of selected topics.
        """
        scorers_topics = {}
        if not hasattr(self, "scorer_topic_vars"):
            raise AttributeError("Topic assignments not found — please build the Scorer–Topic table first.")
    
        for scorer, topics in self.scorer_topic_vars.items():
            selected = [topic for topic, var in topics.items() if var.get()]
            scorers_topics[scorer] = selected
        return scorers_topics
    
    def get_general_topics(self):
        """Return list of general topics checked by user."""
        if not hasattr(self, "general_topic_vars"):
            return []
        return [t for t, var in self.general_topic_vars.items() if var.get()]
    
    def get_scorecard_columns(self):
        """Return list of columns configured for scorecards."""
        if not hasattr(self, "scorecard_cols_listbox"):
            return []
        return [self.scorecard_cols_listbox.get(i) for i in range(self.scorecard_cols_listbox.size())]



    # =========================
    # ADMIN FOLDER SELECTION
    # =========================
    def select_admin_folder(self):
        folder = filedialog.askdirectory(title="Select Admin View Folder")
        if folder:
            self.admin_folder = folder
            self.admin_folder_label.config(text=folder, foreground="black")


    
    def _bind_mousewheel(self, canvas):
        """Enable mousewheel scrolling for a given canvas."""
        def _on_mousewheel(event):
            if canvas.winfo_exists():  # ✅ only scroll if still alive
                canvas.yview_scroll(-1 * int(event.delta / 120), "units")
    
        # Save both handler and canvas for safe cleanup
        self._current_mousewheel_bind = (canvas, _on_mousewheel)
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
    def _unbind_mousewheel(self):
        """Safely remove any previous mousewheel binding."""
        if hasattr(self, "_current_mousewheel_bind"):
            canvas, handler = self._current_mousewheel_bind
            try:
                canvas.unbind_all("<MouseWheel>")
            except Exception:
                pass
            self._current_mousewheel_bind = None

    # =========================
    # HELPER FUNCTIONS FOR LISTBOX
    # =========================
    def move_items(self, from_list, to_list):
        items = [from_list.get(i) for i in from_list.curselection()]
        existing = to_list.get(0, tk.END)
        for item in items:
            if item not in existing:
                to_list.insert(tk.END, item)
        for i in reversed(from_list.curselection()):
            from_list.delete(i)

    def add_custom_option(self, entry, listbox):
        val = entry.get().strip()
        if val and val not in listbox.get(0, tk.END):
            listbox.insert(tk.END, val)
        entry.delete(0, tk.END)

    def remove_selected(self, listbox):
        for i in reversed(listbox.curselection()):
            listbox.delete(i)

    
    def build_dual_option_selector(self, parent, defaults):
        """Two listboxes: available + selected, with add/remove and free entry.
        This frame centers itself inside the given parent using pack(anchor='center')."""
        frame = ttk.Frame(parent)
        frame.pack(pady=5, anchor="center")
    
        # Labels row, centered
        left_label = ttk.Label(frame, text="Available options", justify="center")
        left_label.grid(row=0, column=0, padx=10, pady=(0,6), sticky="nsew")
        right_label = ttk.Label(frame, text="Selected options", justify="center")
        right_label.grid(row=0, column=2, padx=10, pady=(0,6), sticky="nsew")
    
        # Listboxes with centered text
        left_list = tk.Listbox(frame, selectmode=tk.MULTIPLE, height=8, width=40, justify="center")
        right_list = tk.Listbox(frame, selectmode=tk.MULTIPLE, height=8, width=40, justify="center")
        left_list.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")
        right_list.grid(row=1, column=2, padx=10, pady=5, sticky="nsew")
    
        # Populate left_list
        for opt in defaults:
            left_list.insert(tk.END, opt)
    
        # Buttons between lists (centered vertically)
        btns = ttk.Frame(frame)
        btns.grid(row=1, column=1, padx=5, sticky="ns")
        ttk.Button(btns, text="→", width=4, command=lambda: self.move_items(left_list, right_list)).pack(pady=4)
        ttk.Button(btns, text="←", width=4, command=lambda: self.move_items(right_list, left_list)).pack(pady=4)
    
        # Free text add/remove controls under both lists, centered
        entry_frame = ttk.Frame(frame)
        entry_frame.grid(row=2, column=0, columnspan=3, pady=8)
    
        entry = ttk.Entry(entry_frame, width=50, justify="center")
        entry.grid(row=0, column=0, padx=5)
        ttk.Button(entry_frame, text="Add Custom Option → Selected",
                   command=lambda: self.add_custom_option(entry, right_list)).grid(row=0, column=1, padx=5)
        ttk.Button(entry_frame, text="Remove Selected",
                   command=lambda: self.remove_selected(right_list)).grid(row=0, column=2, padx=5)
    
        # Center columns evenly
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_columnconfigure(1, weight=0)
        frame.grid_columnconfigure(2, weight=1)
    
        return (left_list, right_list)

    # =======================================================
    # UNIVERSAL SCORERS IMPORT AND LOAD
    # =======================================================
    def import_scorers_table(self):
        """Import an Excel or CSV scorers file for either simple or topic scoring."""
        filepath = filedialog.askopenfilename(
            title="Select Scorers Table",
            filetypes=[("Excel or CSV files", "*.xlsx *.xls *.csv")]
        )
        if not filepath:
            return
    
        try:
            # Read the file
            if filepath.endswith((".xlsx", ".xls")):
                df = pd.read_excel(filepath)
            else:
                df = pd.read_csv(filepath)
    
            if df.empty:
                messagebox.showwarning("No data", "The selected file appears to be empty.")
                return
    
            # Store file and data
            self.scorers_file = filepath
            self.scorers_df = df
    
            # Update label
            if hasattr(self, "scorers_table_label"):
                self.scorers_table_label.config(text=os.path.basename(filepath), foreground="black")
    
            # Populate dropdown with column names (for user to choose the name column)
            cols = list(df.columns)
            if hasattr(self, "scorer_column_dropdown"):
                self.scorer_column_dropdown["values"] = cols
                self.scorer_column_dropdown.config(state="readonly")
                if hasattr(self, "load_scorers_button"):
                    self.load_scorers_button.config(state="normal")
    
            # Notify user
            messagebox.showinfo(
                "File loaded",
                f"Select the column that contains the scorers’ names."
            )
    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read scorers table:\n{e}")

    
    def load_scorers_from_dropdown(self):
        """Extract scorers' names from the selected column of the imported table."""
        if getattr(self, "scorers_df", None) is None:
            messagebox.showwarning("No file", "Please import a scorers table first.")
            return
    
        if not hasattr(self, "scorer_column_dropdown"):
            messagebox.showerror("UI missing", "Scorer column dropdown not found.")
            return
    
        col = self.scorer_column_dropdown.get()
        if not col:
            messagebox.showwarning("No column selected", "Please select the column containing the scorer names.")
            return
    
        names = (
            self.scorers_df[col]
            .dropna()
            .astype(str)
            .map(str.strip)
            .unique()
            .tolist()
        )
    
        # Add them to whichever scorers_listbox is active
        if hasattr(self, "scorers_listbox"):
            existing = set(self.scorers_listbox.get(0, tk.END))
            for name in names:
                if name not in existing:
                    self.scorers_listbox.insert(tk.END, name)
    
        messagebox.showinfo("Scorers Added", f"✅ {len(names)} scorers added from '{col}'.")

    
    def move_items(self, from_list, to_list):
        """Move selected items between lists."""
        items = [from_list.get(i) for i in from_list.curselection()]
        existing = to_list.get(0, tk.END)
        for item in items:
            if item not in existing:
                to_list.insert(tk.END, item)
        for i in reversed(from_list.curselection()):
            from_list.delete(i)
    
    def add_custom_option(self, entry, listbox):
        val = entry.get().strip()
        if val and val not in listbox.get(0, tk.END):
            listbox.insert(tk.END, val)
        entry.delete(0, tk.END)
    
    def remove_selected(self, listbox):
        for i in reversed(listbox.curselection()):
            listbox.delete(i)
    
    def generate_marking_sheet_from_summary(self):
        """Create Eligibility Marking Sheet directly from the Application Summary.xlsx file."""
    
    
        # --- Check both inputs are selected ---
        if not hasattr(self, "summary_path") or not os.path.exists(self.summary_path):
            messagebox.showwarning("Missing file", "Please select the Application Summary file first.")
            return
    
        if not hasattr(self, "marking_save_path"):
            messagebox.showwarning("Missing save location", "Please choose where to save the Eligibility Marking Sheet.")
            return
    
        # --- Read summary file ---
        try:
            df = pd.read_excel(self.summary_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read Application Summary file:\n{e}")
            return
    
        # --- Validate structure ---
        if df.shape[1] < 2:
            messagebox.showerror(
                "Invalid File",
                "The Application Summary file must contain at least two columns:\n"
                "1. Unique Application ID\n2. Group Name\n(Optional) Topics"
            )
            return
    
        # --- Extract columns safely ---
        id_col = df.columns[0]
        name_col = df.columns[1]
        topics_col = None
        if len(df.columns) > 2:
            topics_col = df.columns[2]
    
        # --- Create group data tuples ---
        group_data = []
        for _, row in df.iterrows():
            uid = str(row[id_col]).strip()
            name = str(row[name_col]).strip()
            topic = str(row[topics_col]).strip() if topics_col else ""
            group_data.append((uid, name, topic))
    
        # --- Create Excel workbook ---
        try:
            self.create_marking_excel_from_summary(group_data, self.marking_save_path)
        except PermissionError:
            messagebox.showerror(
                "File in Use",
                "❌ The Eligibility Marking Sheet could not be saved.\n"
                "Please close the Excel file if it's already open, then try again."
            )
            return
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while saving:\n{e}")
            return
    
        messagebox.showinfo("Success", f"✅ Eligibility Marking Sheet saved")


    def generate_final_decision(self):
        """
        Generate the final eligibility marking summary.
        Uses the selected Eligibility Marking Sheet directly,
        and saves results in the chosen output folder.
        """
        try:
            # Check required inputs
            if not hasattr(self, "final_input_file") or not hasattr(self, "final_output_folder"):
                messagebox.showwarning("Missing Info", "Please select both input file and output folder.")
                return
    
            # Load the marking sheet the user selected
            indata = pd.read_excel(self.final_input_file)
    
            # We no longer need to look for an "Admin View" folder
            # Just pass indata twice — applications can be the same data
            applications = indata.copy()
    
            # Define output file path directly inside chosen output folder
            summary_output = os.path.join(self.final_output_folder, "Eligibility_Marking_Summary.xlsx")
    
            # Run the eligibility marking logic
            eligibility.mark_eligibility(indata, applications, summary_output)
    
    
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while generating the summary:\n{e}")

    
    def generate_scoring_report(self):
        """Generate a Word summary report of scoring setup."""
        try:
            matrix_path = os.path.join(self.scorecards_output_folder, "Matrix.xlsx")
            if not os.path.exists(matrix_path):
                messagebox.showwarning("Missing Matrix", "Please generate the Matrix first.")
                return
            df_matrix = pd.read_excel(matrix_path)
            df_apps = pd.read_excel(self.scorecards_eligible_file)
            from scoring import generate_summary_report
            report_path = generate_summary_report(df_matrix, df_apps, 
                                                  [self.scorers_listbox.get(i) for i in range(self.scorers_listbox.size())],
                                                  int(self.reviews_per_app_var.get()),
                                                  self.scorecards_output_folder)
            messagebox.showinfo("Report Created", f"✅ Report saved at:\n{report_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not generate report:\n{e}")

    

    def select_scorers_folder_for_ranking(self):
        """Let user pick multiple scorer folders under a parent directory, with Select All option."""
        import os
        import tkinter as tk
        from tkinter import filedialog, Toplevel, Checkbutton, IntVar, Label, Button, Frame, ttk, messagebox
    
        # Step 1: Ask for the parent directory
        parent_dir = filedialog.askdirectory(title="Select Parent Folder Containing Scorer Folders")
        if not parent_dir:
            return
    
        # Step 2: Collect subfolders
        subfolders = [
            os.path.join(parent_dir, f)
            for f in os.listdir(parent_dir)
            if os.path.isdir(os.path.join(parent_dir, f))
        ]
    
        if not subfolders:
            messagebox.showwarning("No Folders", "No subfolders found in this directory.")
            return
    
        # Step 3: Create popup window for selection
        top = Toplevel(self.main_frame)
        top.title("Select Scorer Folders")
        top.geometry("420x500")
    
        Label(top, text="Select scorer folders to include:", font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=10, pady=(10, 5)
        )
    
        # Step 4: Frame to hold the scrollable area
        list_frame = Frame(top)
        list_frame.pack(fill="both", expand=True, padx=10, pady=5)
    
        # Scrollable frame setup
        canvas = tk.Canvas(list_frame, borderwidth=0, highlightthickness=0)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        frame = Frame(canvas)
        frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        self._unbind_mousewheel()       # remove any previous binding
        self._bind_mousewheel(canvas)   # bind new one safely

        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
        # Step 5: Variables for each checkbox
        vars = []
        for folder in subfolders:
            var = IntVar()
            chk = Checkbutton(frame, text=os.path.basename(folder), variable=var, anchor="w")
            chk.pack(fill="x", padx=10, pady=2, anchor="w")
            vars.append((var, folder))
    
        # Step 6: Frame for buttons at the bottom
        buttons_frame = Frame(top)
        buttons_frame.pack(fill="x", pady=10)
    
        # Step 7: Add "Select All" checkbox
        select_all_var = IntVar()
    
        def toggle_all():
            state = select_all_var.get()
            for var, _ in vars:
                var.set(state)
    
        select_all_chk = Checkbutton(
            buttons_frame,
            text="Select All",
            variable=select_all_var,
            command=toggle_all,
            font=("Segoe UI", 9, "bold"),
            anchor="w",
        )
        select_all_chk.pack(side="left", padx=20)
    
        # Step 8: Confirm button
        def confirm_selection():
            self.scorer_folders_selected = [folder for var, folder in vars if var.get() == 1]
            count = len(self.scorer_folders_selected)
            folder_names = "\n".join(os.path.basename(f) for f in self.scorer_folders_selected)
            self.scorers_folder_label_ranking.config(
                text=f"📁 {count} folders selected:\n{folder_names}", foreground="black"
            )
            self.scorers_count_label.config(text=f"✅ {count} scorer folder(s) selected.")
            top.destroy()
    
        Button(buttons_frame, text="Confirm Selection", command=confirm_selection).pack(
            side="right", padx=20
        )



            
    def load_responses_excel_for_ranking(self):
        """Load Responses Excel, ask for Organisation Name column, then select columns to include."""
        file_path = filedialog.askopenfilename(
            title="Select Responses Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls")],
        )
        if not file_path:
            return
    
        # --- Safe read ---
        try:
            df = pd.read_excel(file_path)
        except PermissionError:
            messagebox.showerror(
                "File in Use",
                "❌ Please close the Excel file before continuing."
            )
            return
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file:\n{e}")
            return
    
        # --- Save path and show label ---
        self.responses_path = file_path
        self.responses_label.config(
            text=f"✅ Location selected: {os.path.basename(file_path)}",
            foreground="green"
        )
    
        # --- Clear previous widgets ---
        for widget in self.columns_frame.winfo_children():
            widget.destroy()
    
        # === Step 1: Ask which column is the Organisation Name ===
        ttk.Label(
            self.columns_frame,
            text="Select the column that contains the Organisation / Group Name:",
            font=get_font(size=10, weight="bold")
        ).pack(anchor="w", pady=(0, 5))
    
        # Try to auto-detect a likely column
        likely_name_col = None
        for col in df.columns:
            if any(k in col.lower() for k in ["name", "organisation"]):
                likely_name_col = col
                break
        if not likely_name_col:
            likely_name_col = df.columns[0]
    
        # Dropdown selector
        self.org_name_col_var = tk.StringVar(value=likely_name_col)
        name_dropdown = ttk.Combobox(
            self.columns_frame,
            textvariable=self.org_name_col_var,
            values=list(df.columns),
            state="readonly",
            width=60
        )
        name_dropdown.pack(anchor="w", padx=10, pady=(0, 15))
    
        ttk.Label(
            self.columns_frame,
            text="✅ This column will be used to match organisations with the scorecards.",
            foreground="gray"
        ).pack(anchor="w", pady=(0, 10))
    
        # === Step 2: Show checkboxes for columns ===
        ttk.Label(
            self.columns_frame,
            text="Select which columns to include in the final Ranking sheet:",
            font=get_font(size=10, weight="bold")
        ).pack(anchor="w", pady=(10, 5))
    
        wrapper = ttk.Frame(self.columns_frame)
        wrapper.pack(pady=10, padx=40, anchor="center", fill="both", expand=True)
    
        # Scrollable area
        canvas = tk.Canvas(wrapper, height=250, highlightthickness=0)
        scrollbar_y = ttk.Scrollbar(wrapper, orient="vertical", command=canvas.yview)
        scrollbar_x = ttk.Scrollbar(wrapper, orient="horizontal", command=canvas.xview)
    
        scroll_frame = ttk.Frame(canvas, padding=10)
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
    
        canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
    
        # Layout
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
    
        # Checkboxes for each column
        self.selected_columns = {}
        for col in df.columns:
            var = tk.BooleanVar(value=False)
            cb = ttk.Checkbutton(scroll_frame, text=col, variable=var)
            cb.pack(anchor="w", padx=10, pady=2)
            self.selected_columns[col] = var
    
        ttk.Label(
            self.columns_frame,
            text="✅ The selected columns will appear in the final ranking file.",
            foreground="gray"
        ).pack(anchor="w", pady=5)



    
    def run_ranking_process(self):
        """Run the ranking process and create the final ranking Excel file."""
        try:
            from ranking import make_ranking
    
            # --- 1. Check required inputs ---
            if not getattr(self, "scorer_folders_selected", None):
                messagebox.showwarning("Missing Folder", "Please select the scorers folder first.")
                return
    
            if not getattr(self, "responses_path", None):
                messagebox.showwarning("Missing File", "Please load the Responses Excel first.")
                return
    
            if not getattr(self, "output_ranking_path", None):
                messagebox.showwarning("Missing File", "Please choose where to save the ranking file.")
                return
    
            if not hasattr(self, "selected_columns") or not self.selected_columns:
                messagebox.showwarning(
                    "No Columns Selected",
                    "Please load the Responses file and select which columns to include in the ranking."
                )
                return
    
            # --- 2. Read 'how many should pass?' safely ---
            raw_val = ""
            if hasattr(self, "num_pass_var"):
                raw_val = str(self.num_pass_var.get()).strip()
    
            try:
                pass_limit = int(raw_val)
            except (ValueError, TypeError):
                messagebox.showerror(
                    "Invalid Input",
                    "Please enter a valid number for passing applications (e.g. 40)."
                )
                return
    
            # --- 3. Collect selected columns from the Responses checkbox list ---
            selected_cols = [col for col, var in self.selected_columns.items() if var.get()]
            if not selected_cols:
                messagebox.showwarning("No Columns Selected", "Please select at least one column to include.")
                return
    
            # --- 4. Organisation Name column selected in the popup ---
            name_col = self.org_name_col_var.get() if hasattr(self, "org_name_col_var") else None
    
            # --- 5. Run the ranking logic (shading is done inside make_ranking) ---
            make_ranking(
                responses_path=self.responses_path,
                output_ranking=self.output_ranking_path,
                scorers_folders=self.scorer_folders_selected,
                selected_columns=selected_cols,
                top_n_to_shade=pass_limit,
                name_column=name_col,
            )
    
        except Exception as e:
            messagebox.showerror("Ranking Error", f"❌ Error during ranking:\n\n{e}")




    def select_output_file_for_ranking(self):
        """Ask the user where to save the ranking file and show confirmation in green."""
        from tkinter import filedialog
    
        file_path = filedialog.asksaveasfilename(
            title="Select Save Location for Ranking File",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile="Applications Final Ranking.xlsx"
        )
        if not file_path:
            return  # user cancelled
    
        # Save selected path
        self.output_ranking_path = file_path
    
        # Update label in green, like "Load Responses Excel"
        file_name = os.path.basename(file_path)
        self.output_file_label_ranking.config(
            text=f"✅ Location selected: {file_name}",
            foreground="green"
        )



    # -------------------------------
    # Select input file
    # -------------------------------
    def select_file(self):
        """Ask the user to load the main Responses file and populate dropdowns + scorer questions list."""
        filepath = filedialog.askopenfilename(
            title="Select Application Responses File",
            filetypes=[("Excel or CSV files", "*.xlsx *.xls *.csv")]
        )
        if not filepath:
            return
    
        self.filepath = filepath
        self.file_label.config(text=os.path.basename(filepath), fg="black")
    
        # Safely load column headers and populate dropdowns + listbox
        try:
            with safe_open_excel(filepath, "rb") as f:
                df = pd.read_excel(f, nrows=0)
    
            columns = list(df.columns)
            if not columns:
                messagebox.showwarning("Warning", "No columns found in the selected file.")
                return
    
            # --- Populate Organisation Name + Topic dropdowns ---
            self.name_column_dropdown["values"] = columns
            self.name_column_dropdown.config(state="readonly")
            self.topic_column_dropdown["values"] = columns
            self.topic_column_dropdown.config(state="readonly")
    
            # Set default selections (first = org name, second = topic)
            self.name_column_dropdown.current(0)
            if len(columns) > 1:
                self.topic_column_dropdown.current(1)
            else:
                self.topic_column_dropdown.current(0)
    
            # --- Populate Scorer Cards question listbox ---
            if hasattr(self, "column_listbox"):
                self.column_listbox.delete(0, tk.END)
                for col in columns:
                    self.column_listbox.insert(tk.END, col)
                self.column_listbox.selection_clear(0, tk.END)
    
            # Enable generate button now that file is loaded
            if hasattr(self, "generate_button"):
                self.generate_button.config(state="normal")
    
        except PermissionError:
            messagebox.showerror("File in Use", "❌ Please close the file in Excel before selecting it again.")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open the selected file:\n\n{e}")




    # -------------------------------
    # Select output folder
    # -------------------------------
    def select_output_folder(self):
        folder = filedialog.askdirectory(title="Select Output Folder")
        if folder:
            self.output_folder = folder
            self.output_label.config(text=folder, fg="black")

    # -------------------------------
    # Load columns
    # -------------------------------
    def load_columns(self, filepath):
        try:
            if filepath.endswith((".xlsx", ".xls")):
                df = pd.read_excel(filepath, nrows=0)
            else:
                df = pd.read_csv(filepath, nrows=0)
            self.columns = list(df.columns)
    
            for widget in self.columns_frame.winfo_children():
                widget.destroy()

            
            # --- Group name selector ---
            tk.Label(self.columns_frame, text="Select the question showing the organisation name:",
                     font=("Helvetica", 11, "bold")).pack(pady=(10, 0))
            self.group_name_var = tk.StringVar(value=self.columns[0] if self.columns else "")
            self.group_name_dropdown = ttk.Combobox(
                self.columns_frame, textvariable=self.group_name_var,
                values=self.columns, state="readonly", width=80
            )
            self.group_name_dropdown.pack(pady=(0, 10))
            
            # --- NEW: Topics selector (optional) ---
            tk.Label(
                self.columns_frame,
                text="Select the question containing application topics:",
                font=("Helvetica", 11, "bold")
            ).pack(pady=(5, 0))
            self.topic_column_var = tk.StringVar()
            self.topic_column_dropdown = ttk.Combobox(
                self.columns_frame,
                textvariable=self.topic_column_var,
                values=self.columns,
                state="readonly",
                width=80
            )
            self.topic_column_dropdown.pack(pady=(0, 5))
            
            self.no_topics_var = tk.BooleanVar(value=False)
            
            def _toggle_topics_dropdown():
                state = "disabled" if self.no_topics_var.get() else "readonly"
                self.topic_column_dropdown.config(state=state)
            
            ttk.Checkbutton(
                self.columns_frame,
                text="No topics question / skip topics",
                variable=self.no_topics_var,
                command=_toggle_topics_dropdown
            ).pack(pady=(0, 10))
            
            # --- Scorer columns listbox ---
            tk.Label(self.columns_frame, text="Select questions for the Scorer Cards:",
                     font=("Helvetica", 11, "bold")).pack(pady=(10, 0))
            self.column_listbox = tk.Listbox(self.columns_frame, selectmode=tk.MULTIPLE, width=80, height=10)
            for col in self.columns:
                self.column_listbox.insert(tk.END, col)
            self.column_listbox.pack(pady=(5, 10))
            
            # --- ✅ Restore previously selected scorer questions if available ---
            previous_selection = getattr(self, "selected_scorer_questions", [])
            if previous_selection:
                for i, col in enumerate(self.columns):
                    if col in previous_selection:
                        self.column_listbox.selection_set(i)

    
            ttk.Button(self.columns_frame, text="Select All Questions", command=self.select_all_columns).pack(pady=5)
    
            self.generate_button.config(state="normal")
    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read columns from file:\n{e}")

    def select_all_columns(self):
        self.column_listbox.select_set(0, tk.END)
    
        
    
    def load_columns_from_responses(self):
        """Load column names from the selected Responses file for dropdown selection."""
        if not self.filepath:
            messagebox.showwarning("Missing File", "Please select a Responses Excel file first.")
            return
    
        try:
            df = pd.read_excel(self.filepath, nrows=0)
        except Exception as e:
            messagebox.showerror("Error", f"Could not open the Responses file:\n{e}")
            return
    
        cols = list(df.columns)
        self.name_column_dropdown["values"] = cols
        self.name_column_dropdown.config(state="readonly")
        if cols:
            self.name_column_dropdown.current(0)
        messagebox.showinfo("Columns Loaded", f"{len(cols)} columns loaded from Responses file.")


    # -------------------------------
    # Run eligibility creation
    # -------------------------------
    def run_eligibility(self):
        if not self.filepath or not self.output_folder:
            messagebox.showwarning("Missing info", "Please select an Excel/CSV file and output folder.")
            return
    
        # --- Collect selected columns ---
        selected_indices = self.column_listbox.curselection()
        selected_columns = [self.columns[i] for i in selected_indices]
        group_col = self.group_name_var.get()
    
        if not group_col:
            messagebox.showwarning("Missing selection", "Please select the column showing the organisation name.")
            return
    
        if not selected_columns:
            if not messagebox.askyesno("No Columns Selected", "No columns selected for Scorers View. Use ALL columns?"):
                return
            selected_columns = self.columns
    
        # --- Number of applications ---
        num_apps = None
        if hasattr(self, "num_entries_var"):
            try:
                num_apps_str = self.num_entries_var.get().strip()
                num_apps = None if num_apps_str.lower() == "all" else int(num_apps_str)
            except Exception:
                num_apps = None
    
        # --- Topics column ---
        topics_col = None
        if hasattr(self, "no_topics_var") and not self.no_topics_var.get():
            val = self.topic_column_var.get().strip() if hasattr(self, "topic_column_var") else ""
            topics_col = val or None
    
        # --- Create progress popup ---
        self.progress_window = tk.Toplevel(self.root)
        self.progress_window.title("Creating Application Cards...")
        self.progress_window.geometry("400x150")
        tk.Label(self.progress_window, text="Please wait, generating Word files...").pack(pady=10)
        self.progress_bar = ttk.Progressbar(self.progress_window, length=300, mode="determinate")
        self.progress_bar.pack(pady=10)
        self.progress_label = tk.Label(self.progress_window, text="")
        self.progress_label.pack()
    
        # --- Run in background thread ---
        thread = threading.Thread(
            target=self._run_eligibility_background,
            args=(selected_columns, group_col, num_apps, topics_col)
        )
        thread.start()


    def _run_eligibility_background(self, selected_columns, group_col, num_apps, topics_col):
        try:
            create_application_cards(
                self.filepath,
                selected_columns,
                self.output_folder,
                group_col,
                progress_callback=self._update_progress,
                num_apps=num_apps,
                topics_col=topics_col   # ⬅️ pass through
            )
            self._on_task_complete(success=True)
        except Exception as e:
            self._on_task_complete(success=False, error=e)


    def _update_progress(self, current, total):
        percent = int((current / total) * 100)
        self.root.after(0, lambda: self.progress_label.config(text=f"{percent}% completed"))
        self.root.after(0, lambda: self.progress_bar.config(value=percent))


    def _on_task_complete(self, success=True, error=None):
        """Close progress popup automatically when the background task finishes."""
        def close_window():
            # Close the popup safely if it still exists
            if hasattr(self, "progress_window") and self.progress_window.winfo_exists():
                self.progress_window.destroy()
    
        # Schedule the close right away
        self.root.after(0, close_window)
    
        # Show a message once the window is gone
        if success:
            self.root.after(100, lambda: messagebox.showinfo("Success", "✅ Application Cards created successfully!"))
        else:
            self.root.after(100, lambda: messagebox.showerror("Error", f"❌ An error occurred:\n{error}"))
    
    

# -------------------------------
# MAIN ENTRY POINT
# -------------------------------
if __name__ == "__main__":

    # ---------------------------
    # GLOBAL STYLE SETTINGS
    # ---------------------------
    APP_FONT_NAME = "Roboto"  # will fallback if not installed
    BASE_FONT_SIZE = 10

    COLORS = {
        "bg": "#f9f7f4",        # warm light cream background
        "text": "#333333",      # dark text
        "accent": "#FFE5E5",    # super light red (normal state)
        "accent_hover": "#FFB6B6", # soft coral on hover
        "nav": "#0b4d4d",       # teal for navigation
        "nav_hover": "#096363",
        "light_teal": "#b0d1c2" # soft teal highlights
    }

    def setup_global_styles(root):
        """Configure global font and ttk styles."""
        global APP_FONT_NAME
        # Fallback font handling
        try:
            root.tk.call("font", "create", "AppFont", "-family", APP_FONT_NAME, "-size", BASE_FONT_SIZE)
        except tk.TclError:
            print(f"⚠️ Font '{APP_FONT_NAME}' not found. Falling back to Arial.")

            APP_FONT_NAME = "Helvetica"

        # Apply font globally to all Tk widgets
        for f in ("TkDefaultFont", "TkTextFont", "TkMenuFont", "TkHeadingFont"):
            tkfont.nametofont(f).config(family=APP_FONT_NAME, size=BASE_FONT_SIZE)

        # ttk style setup
        style = ttk.Style(root)
        style.theme_use("clam")  # nicer modern base theme

        # Base frame & label background
        root.configure(bg=COLORS["bg"])
        style.configure("TFrame", background=COLORS["bg"])
        style.configure("TLabel", background=COLORS["bg"], foreground=COLORS["text"], font=(APP_FONT_NAME, BASE_FONT_SIZE))
        style.configure("TNotebook", background=COLORS["bg"])
        style.configure("TNotebook.Tab", font=(APP_FONT_NAME, BASE_FONT_SIZE))

        # Buttons
        style.configure("TButton",
                        font=(APP_FONT_NAME, BASE_FONT_SIZE, "bold"),
                        padding=6)

        # Accent (Action) Buttons — e.g. Generate Sheet
        style.configure("Accent.TButton",
                        background=COLORS["accent"],
                        foreground="black",  # better readability
                        borderwidth=0)
        style.map("Accent.TButton",
                  background=[("active", COLORS["accent_hover"])],
                  relief=[("pressed", "sunken")])

        # Navigation Buttons — e.g. Next / Back
        style.configure("Nav.TButton",
                        background=COLORS["nav"],
                        foreground="white",
                        borderwidth=0,
                        focuscolor=COLORS["nav"])
        
        # Navigation Buttons — subtler look for Back / Next
        style.configure(
            "Nav.TButton",
            background="#1c5959",       # softer teal
            foreground="white",
            borderwidth=0,
            font=(APP_FONT_NAME, BASE_FONT_SIZE - 1, "bold"),
            padding=(10, 4),
        )
        style.map(
            "Nav.TButton",
            background=[
                ("active", "#0b4d4d"),   # darken on hover
                ("pressed", "#063636"),  # darker when clicked
            ],
            relief=[("pressed", "sunken")]
            )
        return style

    # ---------------------------
    # SPLASH SCREEN
    # ---------------------------
    splash = tk.Tk()
    splash.overrideredirect(True)
    splash.geometry("300x200+300+300")
    splash.configure(bg=COLORS["bg"])

    logo_path = resource_path("assets/splash.png")
    if os.path.exists(logo_path):
        img = Image.open(logo_path).resize((100, 100), Image.LANCZOS)
        logo = ImageTk.PhotoImage(img)
        tk.Label(splash, image=logo, bg=COLORS["bg"]).pack(pady=10)
        splash.image = logo

    tk.Label(splash, text="Loading Edgy...", font=(APP_FONT_NAME, 12, "bold"), bg=COLORS["bg"], fg=COLORS["nav"]).pack(expand=True)

    def show_splash():
        splash = tk.Tk()
        splash.overrideredirect(True)  # remove title bar
        splash.geometry("300x300")     # square base for circular appearance
        splash.config(bg="#d9f0f0")    # soft teal background
    
        # Make it circular (Windows & macOS work best)
        try:
            splash.wm_attributes("-transparentcolor", "#d9f0f0")  # transparent outside the circle
        except tk.TclError:
            pass  # Fallback: will just appear as rounded-color window
    
        canvas = tk.Canvas(
            splash,
            width=300,
            height=300,
            bg="#d9f0f0",
            highlightthickness=0
        )
        canvas.pack(fill="both", expand=True)
    
        # Draw circular background (white center)
        canvas.create_oval(10, 10, 290, 290, fill="white", outline="")
    
        # Add logo (optional)
        logo_path = resource_path(os.path.join("assets", "logo.png"))
        if os.path.exists(logo_path):
            logo_img = Image.open(logo_path).resize((120, 120), Image.LANCZOS)
            logo = ImageTk.PhotoImage(logo_img)
            canvas.create_image(150, 120, image=logo)
            canvas.image = logo
    
        # Add text
        canvas.create_text(
            150, 220,
            text="Loading Edgy...",
            font=("Helvetica", 12, "bold"),
            fill="#333333"
        )
    
        # Center splash on screen
        splash.update_idletasks()
        w, h = 100, 100
        ws = splash.winfo_screenwidth()
        hs = splash.winfo_screenheight()
        x = int((ws/2) - (w/2))
        y = int((hs/2) - (h/2))
        splash.geometry(f"{w}x{h}+{x}+{y}")

    def start_main_app():
        splash.destroy()
        root = tk.Tk()
        root.title("Edge Fund Application Toolkit")
        root.geometry("850x600")
        root.resizable(False, False)

        setup_global_styles(root)

        png_icon_path = resource_path(os.path.join("assets", "logo.png"))
        ico_icon_path = resource_path(os.path.join("assets", "logo.ico"))
        if os.path.exists(png_icon_path):
            icon_img_main = ImageTk.PhotoImage(Image.open(png_icon_path))
            root.iconphoto(True, icon_img_main)
        if os.path.exists(ico_icon_path):
            root.iconbitmap(ico_icon_path)

        app = EdgeFundApp(root)
        root.mainloop()

    splash.after(2500, start_main_app)
    splash.mainloop()


